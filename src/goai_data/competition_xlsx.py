"""Competition-style XLSX export and import for simulated GoAI matches.

The visible worksheets follow the Chinese enterprise competition exports.  A
small hidden metadata sheet preserves provenance and rule-pack identity, while
events, orders, reports and final cash are imported from the visible tables.
"""

from __future__ import annotations

import json
import tempfile
import zipfile
from io import BytesIO
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .common import as_int, as_number, clean_text, parse_money_wan, parse_period
from .full_sandbox import FULL_SANDBOX_VERSION, FinancialSandboxState, FullCompetitionArena


COMPETITION_XLSX_VERSION = "goai_competition_xlsx_v1.0"
COMPETITION_ARCHIVE_VERSION = "stratpilot_competition_archive_v1.0"
ENTERPRISE_SHEETS = (
    "企业信息",
    "库存信息",
    "银行贷款",
    "研发认证",
    "厂房与生产线",
    "订单信息",
    "现金流量表",
    "三张报表",
    "广告投放",
)

EVENT_ACTIONS = {
    "factory_purchased": "Buy_Workshop",
    "factory_rented": "Rent_Workshop",
    "factory_sold": "Sell_Workshop",
    "production_line_ordered": "Buy_ProductLine",
    "production_line_sold": "Sell_ProductLine",
    "production_line_converted": "Transfer",
    "material_ordered": "Buy_Material",
    "material_received": "Pay_Material",
    "emergency_material_purchase": "Emergency_Buy_Material",
    "production_started": "Product_Produce",
    "product_development_expense": "Develop_Product",
    "market_development_expense": "Develop_Market",
    "iso_development_expense": "Develop_ISO",
    "advertising_expense": "Pay_AD",
    "auction_bid_fee": "Auction_Win",
    "order_delivered": "Sell_Product",
    "order_default_penalty": "Pay_Punish",
    "receivable_collected": "Update_Receivable",
    "receivable_discounted": "Discount",
    "short_loan_borrowed": "Short_Loan",
    "short_loan_repaid": "Pay_Short_Loan",
    "short_loan_interest_paid": "Pay_Short_Loan",
    "long_loan_borrowed": "Long_Loan",
    "long_loan_repaid": "Pay_Long_Loan",
    "long_loan_interest_paid": "Pay_Long_Loan",
    "management_fee_expense": "Pay_Overhaul",
    "maintenance_expense": "Pay_Maintenance",
    "factory_rent_expense": "Renew_Workshop",
    "tax_expense": "Pay_Tax",
    "depreciation": "Depreciation",
}

ACTION_EVENTS = {value: key for key, value in EVENT_ACTIONS.items()}


def _json(path: Path, value: Any) -> None:
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def _jsonl(path: Path, rows: Iterable[Mapping[str, Any]]) -> None:
    path.write_text("".join(json.dumps(dict(row), ensure_ascii=False, sort_keys=True) + "\n" for row in rows), encoding="utf-8")


def _money(value: Any) -> str:
    number = float(value or 0)
    return f"{number:.12g}W"


def _period_cn(period: Any, *, year_only: bool = False) -> str:
    text = str(period or "")
    if text.startswith("Y"):
        year_text, _, quarter_text = text[1:].partition("Q")
        if year_text.isdigit():
            if year_only or not quarter_text:
                return f"第{int(year_text)}年"
            if quarter_text.isdigit():
                return f"第{int(year_text)}年{int(quarter_text)}季"
    return text or "-"


def _period_index_cn(period_index: Any) -> str:
    """Render a zero-based quarter index using the reference workbook style."""
    if period_index is None or period_index == "":
        return "-"
    try:
        index = int(period_index)
    except (TypeError, ValueError):
        return _period_cn(period_index)
    return f"第{index // 4 + 1}年{index % 4 + 1}季" if index >= 0 else "-"


def _header(ws, row: int, values: Sequence[Any], start_col: int = 2) -> None:
    fill = PatternFill("solid", fgColor="D9EAF7")
    for offset, value in enumerate(values):
        cell = ws.cell(row, start_col + offset, value)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _title(ws, text: str, *, row: int = 2, start_col: int = 2, end_col: int = 8) -> None:
    ws.cell(row, start_col, text)
    ws.cell(row, start_col).font = Font(bold=True, size=14)
    if end_col > start_col:
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)


def _finish_sheet(ws, widths: Mapping[int, float] | None = None) -> None:
    ws.freeze_panes = "B4"
    ws.sheet_view.showGridLines = False
    for index, width in (widths or {}).items():
        ws.column_dimensions[get_column_letter(index)].width = width


def _metadata_sheet(workbook: Workbook, values: Mapping[str, Any]) -> None:
    ws = workbook.create_sheet("_GOAI_META")
    ws.sheet_state = "hidden"
    ws.append(["key", "value"])
    for key, value in values.items():
        ws.append([key, json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value])


def _merge_global_orders(orders: Sequence[Mapping[str, Any]], arena: FullCompetitionArena) -> list[dict[str, Any]]:
    ownership: dict[str, dict[str, Any]] = {}
    for state in arena.states.values():
        for order in state.assigned_orders:
            ownership[str(order["order_id"])] = dict(order)
    output = []
    for source in orders:
        row = dict(source)
        assigned = ownership.get(str(row.get("order_id")))
        if assigned:
            row.update(assigned)
        output.append(row)
    return output


def _write_enterprise_info(workbook: Workbook, state: FinancialSandboxState) -> None:
    ws = workbook.active
    ws.title = "企业信息"
    _title(ws, f"{state.team_id}公司详细资料", end_col=5)
    ws["B3"], ws["C3"] = "制表人", "GoAI 模拟器"
    ws["D3"], ws["E3"] = "制表时间", datetime.now().astimezone().isoformat(timespec="seconds")
    ws["B4"], ws["C4"] = "公司现金", _money(state.cash_wan)
    ws["D4"], ws["E4"] = "公司状态", "破产" if state.bankrupt else "经营结束"
    ws["B5"], ws["C5"] = "股东注资", "0W"
    ws["D5"], ws["E5"] = "系统时间", "第6年1季"
    ws["B6"], ws["C6"] = "公司名称", f"{state.team_id}模拟公司"
    ws["D6"], ws["E6"] = "所属学校", "模拟赛场"
    ws["B7"], ws["C7"] = "组织结构", "总经理:Agent 财务总监:Agent 市场总监:Agent 运营总监:Agent"
    ws["B8"], ws["C8"] = "公司宣言", "simulated competition export"
    _finish_sheet(ws, {2: 18, 3: 35, 4: 18, 5: 28})


def _write_inventory(workbook: Workbook, state: FinancialSandboxState) -> None:
    ws = workbook.create_sheet("库存信息")
    ws["B2"], ws["G2"], ws["J2"] = "原料订购", "原料库存", "产品库存"
    _header(ws, 3, ["名称", "数量", "剩余时间", "订购时间"], 2)
    _header(ws, 3, ["名称", "数量"], 7)
    _header(ws, 3, ["名称", "数量"], 10)
    row = 4
    for pending in state.pending_material_orders:
        ws.cell(row, 2, pending.get("material_id")); ws.cell(row, 3, pending.get("quantity")); ws.cell(row, 4, f"{max(0, int(pending.get('arrival_period_index', state.period_index)) - state.period_index)}季"); ws.cell(row, 5, _period_cn(pending.get("ordered_period")))
        row += 1
    for index, (name, quantity) in enumerate(state.material_inventory.items(), 4):
        ws.cell(index, 7, name); ws.cell(index, 8, quantity)
    for index, (name, quantity) in enumerate(state.product_inventory.items(), 4):
        ws.cell(index, 10, name); ws.cell(index, 11, quantity)
    _finish_sheet(ws, {2: 14, 3: 12, 4: 14, 5: 18, 7: 14, 8: 12, 10: 14, 11: 12})


def _write_loans(workbook: Workbook, state: FinancialSandboxState) -> None:
    ws = workbook.create_sheet("银行贷款")
    ws["B2"], ws["E2"], ws["I2"], ws["M2"] = "应收款", "长期贷款", "短期贷款", "特别贷款"
    _header(ws, 3, ["剩余账期", "金额"], 2)
    _header(ws, 3, ["剩余账期", "金额", "贷款时间"], 5)
    _header(ws, 3, ["剩余账期", "金额", "贷款时间"], 9)
    _header(ws, 3, ["金额", "贷款时间"], 13)
    for index, item in enumerate(state.receivables, 4):
        ws.cell(index, 2, f"{max(0, int(item.get('due_period_index', state.period_index)) - state.period_index)}季"); ws.cell(index, 3, _money(item.get("amount_wan")))
    for index, item in enumerate(state.long_loans, 4):
        ws.cell(index, 5, f"{max(0, int(item.get('due_period_index', state.period_index)) - state.period_index) / 4:g}年"); ws.cell(index, 6, _money(item.get("principal_wan"))); ws.cell(index, 7, _period_index_cn(item.get("borrowed_period_index")))
    for index, item in enumerate(state.short_loans, 4):
        ws.cell(index, 9, f"{max(0, int(item.get('due_period_index', state.period_index)) - state.period_index)}季"); ws.cell(index, 10, _money(item.get("principal_wan"))); ws.cell(index, 11, _period_index_cn(item.get("borrowed_period_index")))
    _finish_sheet(ws, {2: 14, 3: 14, 5: 14, 6: 14, 7: 18, 9: 14, 10: 14, 11: 18, 13: 14, 14: 18})


def _write_development(workbook: Workbook, state: FinancialSandboxState, rules: Mapping[str, Any]) -> None:
    ws = workbook.create_sheet("研发认证")
    params = rules.get("parameters", rules)
    ws["B2"], ws["H2"], ws["O2"] = "市场开拓", "产品研发", "ISO认证"
    headings = ["名称", "开发费", "周期", "开始时间", "剩余时间", "完成时间"]
    _header(ws, 3, headings, 2); _header(ws, 3, headings, 8); _header(ws, 3, headings, 15)
    completed = {(str(row.get("kind")), str(row.get("target"))): row for row in state.completed_development}
    pending = {(str(row.get("kind")), str(row.get("target"))): row for row in state.pending_development}
    for index, (name, value) in enumerate((params.get("markets") or {}).items(), 4):
        record, waiting = completed.get(("market", str(name))), pending.get(("market", str(name)))
        values = [name, f"{value.get('fee_wan_per_year', 0):g}W/年", f"{value.get('years', 0):g}年", _period_cn(waiting.get("started_period")) if waiting else (_period_cn(record.get("started_period")) if record else "初始资格"), f"{waiting.get('remaining_installments', 0)}次" if waiting else "-", _period_cn(record.get("completed_period")) if record else "-"]
        for col, cell_value in enumerate(values, 2): ws.cell(index, col, cell_value)
    for index, (name, value) in enumerate((params.get("products") or {}).items(), 4):
        record, waiting = completed.get(("product", str(name))), pending.get(("product", str(name)))
        values = [name, f"{value.get('development_wan_per_quarter', 0):g}W/季", f"{value.get('quarters', 0):g}季", _period_cn(waiting.get("started_period")) if waiting else (_period_cn(record.get("started_period")) if record else "初始资格"), f"{waiting.get('remaining_installments', 0)}季" if waiting else "-", _period_cn(record.get("completed_period")) if record else "-"]
        for col, cell_value in enumerate(values, 8): ws.cell(index, col, cell_value)
    for index, (name, value) in enumerate((params.get("iso") or {}).items(), 4):
        record, waiting = completed.get(("iso", str(name))), pending.get(("iso", str(name)))
        values = [name, f"{value.get('fee_wan_per_year', 0):g}W/年", f"{value.get('years', 0):g}年", _period_cn(waiting.get("started_period")) if waiting else (_period_cn(record.get("started_period")) if record else "初始资格"), f"{waiting.get('remaining_installments', 0)}年" if waiting else "-", _period_cn(record.get("completed_period")) if record else "-"]
        for col, cell_value in enumerate(values, 15): ws.cell(index, col, cell_value)
    _finish_sheet(ws, {index: 15 for index in range(2, 21)})


def _write_assets(workbook: Workbook, state: FinancialSandboxState, rules: Mapping[str, Any]) -> None:
    ws = workbook.create_sheet("厂房与生产线")
    params = rules.get("parameters", rules)
    ws["B2"] = "厂房信息"
    _header(ws, 3, ["ID", "名称", "状态", "容量", "购价", "租金", "售价", "最后付租", "置办时间"], 2)
    row = 4
    for item in state.factories:
        spec = (params.get("factories") or {}).get(item.get("name"), {})
        acquired = item.get("acquired_period") or _period_index_cn(item.get("acquired_period_index"))
        values = [item.get("factory_id"), item.get("name"), "购买" if item.get("ownership") == "purchased" else "租用", f"0/{item.get('capacity', 0)}", _money(item.get("cost_wan", spec.get("purchase_wan"))), f"{spec.get('rent_wan_per_year', 0):g}W/年", _money(item.get("book_value_wan")), "-", _period_cn(acquired)]
        for col, value in enumerate(values, 2): ws.cell(row, col, value)
        row += 1
    row += 1
    ws.cell(row, 2, "生产线信息")
    _header(ws, row + 1, ["ID", "名称", "厂房", "产品", "状态", "累计折旧", "开产时间", "转产时间", "剩余时间", "建成时间", "开建时间"], 2)
    row += 2
    for item in state.production_lines:
        completed = item.get("completed_period") or _period_index_cn(item.get("completion_period_index"))
        ordered = item.get("ordered_period") or _period_index_cn(item.get("ordered_period_index"))
        values = [item.get("line_id"), item.get("line_type"), item.get("factory_id") or "-", item.get("product_id") or "-", item.get("status"), _money(item.get("accumulated_depreciation_wan")), "-", "-", "0季", _period_cn(completed), _period_cn(ordered)]
        for col, value in enumerate(values, 2): ws.cell(row, col, value)
        row += 1
    _finish_sheet(ws, {index: 16 for index in range(2, 13)})


def _write_orders(workbook: Workbook, state: FinancialSandboxState) -> None:
    ws = workbook.create_sheet("订单信息")
    _title(ws, "订单列表", end_col=12)
    _header(ws, 3, ["订单编号", "市场", "产品", "数量", "总价", "状态", "得单年份", "交货期", "账期", "ISO", "交货时间"], 2)
    for row, order in enumerate(sorted(state.assigned_orders, key=lambda item: str(item.get("order_id"))), 4):
        values = [order.get("order_id"), order.get("market"), order.get("product"), order.get("quantity"), _money(order.get("total_price_wan")), order.get("status"), _period_cn(order.get("year"), year_only=True) if str(order.get("year", "")).startswith("Y") else f"第{order.get('year')}年", f"{order.get('delivery_term_quarters', 0)}季", f"{order.get('receivable_term_quarters', 0)}季", order.get("iso") or "-", _period_cn(order.get("delivered_period"))]
        for col, value in enumerate(values, 2): ws.cell(row, col, value)
    _finish_sheet(ws, {2: 24, 3: 12, 4: 10, 5: 10, 6: 12, 7: 12, 8: 14, 9: 12, 10: 12, 11: 14, 12: 16})


def _event_note(event: Mapping[str, Any]) -> str:
    """Translate journal events to human-readable competition bookkeeping notes."""
    event_type = str(event.get("event_type") or "")
    details = event.get("details") if isinstance(event.get("details"), Mapping) else event
    labels = {
        "factory_purchased": "购买厂房", "factory_rent_expense": "支付厂房租金", "factory_sold": "出售厂房",
        "production_line_ordered": "订购生产线", "production_line_investment": "支付生产线投资款",
        "production_line_ready": "生产线安装完成", "production_line_sold": "出售生产线",
        "production_line_conversion_expense": "生产线转产", "material_ordered": "订购原料",
        "material_arrived": "支付原料费", "emergency_material_purchase": "紧急采购原料",
        "emergency_product_purchase": "紧急采购产品", "production_started": "开始生产",
        "production_completed": "生产完成", "order_delivered": "交付订单并确认销售收入",
        "order_default_penalty": "支付订单违约损失", "receivable_collected": "收回应收款",
        "receivable_discounted": "贴现应收款", "short_loan_borrowed": "申请短期贷款",
        "short_loan_repaid": "偿还短期贷款本息", "long_loan_borrowed": "申请长期贷款",
        "long_loan_principal_repaid": "偿还长期贷款本金", "long_loan_interest_paid": "支付长期贷款利息",
        "management_fee_expense": "支付行政管理费", "maintenance_expense": "支付生产线维护费",
        "advertising_expense": "广告投放", "auction_bid_fee": "支付竞单费用",
        "spy_information_purchase": "购买竞争情报", "tax_expense": "计提所得税",
        "income_tax_expense": "计提所得税", "income_tax_paid": "缴纳所得税",
        "depreciation": "计提折旧", "development_completed": "研发或认证完成",
        "product_development_expense": "产品生产资格投资", "market_development_expense": "市场开拓资格投资",
        "iso_development_expense": "ISO认证资格投资",
    }
    note = labels.get(event_type, event_type or "经营事项")
    if event_type in {"factory_purchased", "factory_rent_expense", "factory_sold"}:
        factory = details.get("factory") if isinstance(details.get("factory"), Mapping) else details
        name = factory.get("name") if isinstance(factory, Mapping) else details.get("factory")
        if name:
            note += f"[{name}]"
    elif event_type == "production_line_ordered":
        line = details.get("line") if isinstance(details.get("line"), Mapping) else details
        if isinstance(line, Mapping) and line.get("line_type"):
            note += f"[{line.get('line_type')}]"
        if isinstance(line, Mapping) and line.get("product_id"):
            note += f"生产[{line.get('product_id')}]"
    elif event_type in {"material_ordered", "material_arrived"}:
        rows = details.get("orders") if isinstance(details.get("orders"), list) else ([details.get("order")] if isinstance(details.get("order"), Mapping) else [])
        parts = [f"{row.get('material_id')}:{float(row.get('quantity')):g}" for row in rows if isinstance(row, Mapping) and row.get("material_id") is not None]
        if parts:
            note += " " + " ".join(parts)
    elif event_type == "production_line_conversion_expense" and details.get("product_id"):
        note += f"，转产成{details.get('product_id')}"
    elif event_type in {"production_started", "production_completed"}:
        job = details.get("job") if isinstance(details.get("job"), Mapping) else details
        if isinstance(job, Mapping) and job.get("product_id"):
            note += f"[{job.get('product_id')}]"
    elif event_type in {"order_delivered", "order_default_penalty"} and details.get("order_id"):
        note += f"，订单{details.get('order_id')}"
    elif event_type == "development_completed" and details.get("target"):
        note += f"：{details.get('target')}"
    return note


def _write_cashflow(workbook: Workbook, state: FinancialSandboxState, initial_cash: float) -> None:
    ws = workbook.create_sheet("现金流量表")
    _title(ws, "现金流量表", end_col=7)
    _header(ws, 3, ["ID", "动作", "资金", "余额", "时间", "备注"], 2)
    balance = float(initial_cash)
    ws.append([])
    values = [1, "Pay_Capital", initial_cash, balance, "第1年1季", "公司成立，股东注资"]
    for col, value in enumerate(values, 2): ws.cell(4, col, value)
    for sequence, event in enumerate(state.journal, 2):
        amount = float(event.get("cash_effect_wan") or 0)
        balance += amount
        row = sequence + 3
        values = [sequence, EVENT_ACTIONS.get(str(event.get("event_type")), str(event.get("event_type"))), amount, balance, _period_cn(event.get("period")), _event_note(event)]
        for col, value in enumerate(values, 2): ws.cell(row, col, value)
    _finish_sheet(ws, {2: 10, 3: 28, 4: 14, 5: 14, 6: 16, 7: 80})


EXPENSE_METRICS = ("管理费", "广告费", "维护费", "损失", "转产费", "租金", "市场开拓费", "产品研发费", "ISO认证费", "信息费", "合计")
INCOME_METRICS = ("销售收入", "直接成本", "毛利", "综合费用", "折旧前利润", "折旧", "支付利息前利润", "财务费用", "税前利润", "所得税", "年度净利润")
BALANCE_METRICS = ("现金", "应收款", "在制品", "产成品", "原料", "流动资产合计", "厂房", "机器设备", "在建工程", "固定资产合计", "资产总计", "长期贷款", "短期贷款", "特别贷款", "所得税", "负债合计", "股东资本", "利润留存", "年度净利", "所有者权益合计", "负债和所有者权益总计")


def _report_values(report: Mapping[str, Any], initial_cash: float) -> tuple[dict[str, float], dict[str, float], dict[str, float]]:
    income = report.get("income_statement") or {}
    details = income.get("details") or {}
    cash_details = (report.get("cash_flow_statement") or {}).get("details") or {}
    expense = {
        "管理费": abs(float(details.get("management_fee_expense", 0))),
        "广告费": abs(float(details.get("advertising_expense", 0))),
        "维护费": abs(float(details.get("maintenance_expense", 0))),
        "损失": abs(float(details.get("order_default_penalty", 0))),
        "转产费": abs(float(details.get("production_line_conversion", 0))),
        "租金": abs(float(details.get("factory_rent_expense", 0))),
        "市场开拓费": abs(float(details.get("market_development_expense", 0))),
        "产品研发费": abs(float(details.get("product_development_expense", 0))),
        "ISO认证费": abs(float(details.get("iso_development_expense", 0))),
        "信息费": abs(float(details.get("spy_information_purchase", 0))),
    }
    expense["合计"] = sum(expense.values())
    revenue = float(income.get("revenue_wan", 0)); direct = abs(float(income.get("cost_of_goods_sold_wan", 0))); depreciation = abs(float(details.get("depreciation_expense", 0))); finance = abs(float(details.get("interest_expense", 0))); tax = abs(float(income.get("income_tax_wan", 0))); net = float(income.get("net_income_wan", 0))
    income_values = {"销售收入": revenue, "直接成本": direct, "毛利": revenue - direct, "综合费用": expense["合计"], "折旧前利润": revenue - direct - expense["合计"], "折旧": depreciation, "支付利息前利润": revenue - direct - expense["合计"] - depreciation, "财务费用": finance, "税前利润": net + tax, "所得税": tax, "年度净利润": net}
    balance = report.get("balance_sheet") or {}
    fixed = float(balance.get("fixed_assets_wan", 0)); owner = float(balance.get("owner_equity_wan", 0)); liabilities = float(balance.get("liabilities_wan", 0)); total = float(balance.get("total_assets_wan", 0))
    balance_values = {
        "现金": float(balance.get("cash_wan", 0)), "应收款": float(balance.get("receivables_wan", 0)), "在制品": float(balance.get("work_in_process_wan", 0)), "产成品": float(balance.get("products_wan", 0)), "原料": float(balance.get("materials_wan", 0)), "流动资产合计": float(balance.get("current_assets_wan", 0)), "厂房": 0.0, "机器设备": fixed, "在建工程": 0.0, "固定资产合计": fixed, "资产总计": total, "长期贷款": 0.0, "短期贷款": liabilities, "特别贷款": 0.0, "所得税": 0.0, "负债合计": liabilities, "股东资本": initial_cash, "利润留存": owner - initial_cash - net, "年度净利": net, "所有者权益合计": owner, "负债和所有者权益总计": liabilities + owner,
    }
    return expense, income_values, balance_values


def _write_reports(workbook: Workbook, state: FinancialSandboxState, initial_cash: float) -> None:
    ws = workbook.create_sheet("三张报表")
    periods = ["初始元年"] + [f"第{year}年" for year in range(1, 6)]
    _header(ws, 2, ["年度", *periods], 2)
    for row, metric in enumerate(EXPENSE_METRICS, 3): ws.cell(row, 2, metric)
    _header(ws, 15, ["年度", *periods], 2)
    for row, metric in enumerate(INCOME_METRICS, 16): ws.cell(row, 2, metric)
    _header(ws, 28, ["年度", *periods], 2)
    ws["B29"] = "类型"
    for col in range(3, 9): ws.cell(29, col, "系统")
    for row, metric in enumerate(BALANCE_METRICS, 30): ws.cell(row, 2, metric)
    initial_balance = {metric: 0.0 for metric in BALANCE_METRICS}
    initial_balance.update({"现金": initial_cash, "流动资产合计": initial_cash, "资产总计": initial_cash, "股东资本": initial_cash, "所有者权益合计": initial_cash, "负债和所有者权益总计": initial_cash})
    for row, metric in enumerate(EXPENSE_METRICS, 3): ws.cell(row, 3, 0.0)
    for row, metric in enumerate(INCOME_METRICS, 16): ws.cell(row, 3, 0.0)
    for row, metric in enumerate(BALANCE_METRICS, 30): ws.cell(row, 3, initial_balance[metric])
    by_year = {int(report.get("year")): report for report in state.reports}
    for year in range(1, 6):
        report = by_year.get(year, {})
        expense, income, balance = _report_values(report, initial_cash)
        col = year + 3
        for row, metric in enumerate(EXPENSE_METRICS, 3): ws.cell(row, col, expense.get(metric, 0.0))
        for row, metric in enumerate(INCOME_METRICS, 16): ws.cell(row, col, income.get(metric, 0.0))
        for row, metric in enumerate(BALANCE_METRICS, 30): ws.cell(row, col, balance.get(metric, 0.0))
    _finish_sheet(ws, {2: 24, **{index: 16 for index in range(3, 9)}})


def _write_advertising(workbook: Workbook, state: FinancialSandboxState, rules: Mapping[str, Any]) -> None:
    ws = workbook.create_sheet("广告投放")
    params = rules.get("parameters", rules)
    markets = list((params.get("markets") or {}).keys())
    products = list((params.get("products") or {}).keys())
    row = 2
    for year in range(1, 6):
        ws.cell(row, 2, f"第{year}年广告投放情况")
        _header(ws, row + 1, ["产品", *markets], 2)
        for product_index, product in enumerate(products, row + 2):
            ws.cell(product_index, 2, product)
            for market_index, market in enumerate(markets, 3):
                ws.cell(product_index, market_index, float(state.advertising.get(f"{market}:{product}", 0)))
        row += len(products) + 3
    _finish_sheet(ws, {2: 16, **{index: 14 for index in range(3, 3 + len(markets))}})


def _write_enterprise_workbook(path: Path, state: FinancialSandboxState, rules: Mapping[str, Any]) -> None:
    initial_cash = float((rules.get("parameters") or {}).get("initial_cash_wan", 0))
    workbook = Workbook()
    _write_enterprise_info(workbook, state)
    _write_inventory(workbook, state)
    _write_loans(workbook, state)
    _write_development(workbook, state, rules)
    _write_assets(workbook, state, rules)
    _write_orders(workbook, state)
    _write_cashflow(workbook, state, initial_cash)
    _write_reports(workbook, state, initial_cash)
    _write_advertising(workbook, state, rules)
    _metadata_sheet(workbook, {"format_version": COMPETITION_XLSX_VERSION, "match_id": state.match_id, "team_id": state.team_id, "rule_pack_id": rules.get("rule_pack_id"), "provenance": "simulated"})
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _write_global_orders(path: Path, orders: Sequence[Mapping[str, Any]], rules: Mapping[str, Any]) -> None:
    workbook = Workbook(); ws = workbook.active; ws.title = "订单详情"
    _header(ws, 1, ["订单编号", "类型", "年份", "市场", "产品", "数量", "总价", "交货期", "账期", "ISO", "所属用户", "状态"], 1)
    for row, order in enumerate(orders, 2):
        values = [order.get("order_id"), order.get("order_type"), order.get("year"), order.get("market"), order.get("product"), order.get("quantity"), order.get("total_price_wan"), order.get("delivery_term_quarters"), order.get("receivable_term_quarters"), order.get("iso") or "-", order.get("owner_team_id"), order.get("status")]
        for col, value in enumerate(values, 1): ws.cell(row, col, value)
    _finish_sheet(ws, {1: 24, 2: 12, 3: 10, 4: 12, 5: 10, 6: 10, 7: 12, 8: 12, 9: 12, 10: 14, 11: 18, 12: 14})
    _metadata_sheet(workbook, {"format_version": COMPETITION_XLSX_VERSION, "match_id": rules.get("match_id"), "rule_pack_id": rules.get("rule_pack_id"), "provenance": "simulated"})
    workbook.save(path)


def _write_rules(path: Path, rules: Mapping[str, Any]) -> None:
    workbook = Workbook(); ws = workbook.active; ws.title = "Sheet1"
    params = rules.get("parameters", rules)
    long_loan = params.get("long_loan") or {}
    short_loan = params.get("short_loan") or {}
    discount = params.get("receivable_discount") or {}
    asset_limits = params.get("asset_limits") or {}
    ws.append(["重要经营规则"])
    ws.append([f"模拟器版本:{FULL_SANDBOX_VERSION}"])
    ws.append([f"当前规则方案名称：{rules.get('rule_pack_id')}"])
    ws.append(["一.生产线"])
    ws.append(["名称", "投资总额", "每季投资额", "安装周期", "生产周期", "每季转产费", "转产周期", "维护费", "残值", "折旧费", "折旧时间", "分值"])
    for name, item in (params.get("production_lines") or {}).items():
        ws.append([
            name, _money(item.get("investment_wan")), _money(item.get("investment_wan_per_quarter")),
            f"{item.get('install_quarters', 0)}季", f"{item.get('production_quarters', 0)}季",
            _money(item.get("conversion_wan_per_quarter")), f"{item.get('conversion_quarters', 0)}季",
            f"{item.get('maintenance_wan_per_year', 0):g}W/年", _money(item.get("residual_value_wan")),
            _money(item.get("depreciation_fee_wan")), f"{item.get('depreciation_years', 0)}年", item.get("score"),
        ])
    ws.append(["安装周期为0表示即买即用；只有空闲生产线可以转产或出售。"])
    ws.append(["生产线按平均年限法计提折旧，建成当年不提折旧，净值达到残值后停止折旧。"])
    ws.append(["当年建成的生产线需要按本场规则支付维护费。"])
    ws.append(["二.融资"])
    ws.append(["贷款类型", "贷款时间", "贷款额度", "年息/贴现率", "还款方式", "备注"])
    ws.append(["长期贷款", "每年年初", f"长短贷合计不超过上年权益 {long_loan.get('max_total_multiple_prior_equity', 0)} 倍", long_loan.get("annual_rate"), "按年付息，到期还本", f"最小 {long_loan.get('minimum_wan', 0)}W"])
    ws.append(["短期贷款", "每季度初", "按本企业状态和规则校验", short_loan.get("rate"), "到期一次还本付息", "期限按季度"])
    ws.append(["资金贴现", "季度决策时", "不超过可贴现应收款", discount.get("terms_1_2"), "变现时贴息", f"3、4期贴现率 {discount.get('terms_3_4')}"])
    ws.append(["库存拍卖", "应急阶段", "按库存数量", "-", "即时结算", "产品和原料折价率以本场规则为准"])
    ws.append(["三.厂房"])
    ws.append(["名称", "购买价格", "租用价格", "出售价格", "生产线容量", "使用上限", "分值"])
    for name, item in (params.get("factories") or {}).items():
        ws.append([name, _money(item.get("purchase_wan")), f"{item.get('rent_wan_per_year', 0)}W/年", _money(item.get("sale_wan")), item.get("capacity"), item.get("usage_limit", asset_limits.get("max_factories_per_type")), item.get("score")])
    ws.append([f"本场厂房总数上限为 {asset_limits.get('max_factories_total', params.get('max_factory_count', '-'))}；同类型上限为 {asset_limits.get('max_factories_per_type', '-')}。"])
    ws.append([f"厂房出售形成 {params.get('factory_sale_receivable_term_quarters', 0)} 个账期的应收款；租金和续租由环境按规则结算。"])
    ws.append(["四.市场开拓"]); ws.append(["名称", "每年开发费", "开发时间", "分值"])
    for name, item in (params.get("markets") or {}).items(): ws.append([name, _money(item.get("fee_wan_per_year")), f"{item.get('years', 0)}年", item.get("score")])
    ws.append(["市场开发费用在年末支付，不允许加速；完成后取得对应市场资格。"])
    ws.append(["五.ISO认证"]); ws.append(["名称", "每年开发费", "开发时间", "分值"])
    for name, item in (params.get("iso") or {}).items(): ws.append([name, _money(item.get("fee_wan_per_year")), f"{item.get('years', 0)}年", item.get("score")])
    ws.append(["ISO 开发费用在年末支付，不允许加速；完成后取得对应认证。"])
    ws.append(["六.产品研发"]); ws.append(["名称", "加工费", "每季开发费", "开发时间", "直接成本", "分值", "产品组成"])
    for name, item in (params.get("products") or {}).items(): ws.append([name, _money(item.get("process_wan")), _money(item.get("development_wan_per_quarter")), f"{item.get('quarters', 0)}季", _money(item.get("direct_cost_wan")), item.get("score"), ", ".join(f"{count}*{material}" if count != 1 else material for material, count in (item.get("bom") or {}).items())])
    ws.append(["产品研发费用在季末支付，不允许加速；完成后取得对应生产资格。"])
    ws.append(["七.原料设置"]); ws.append(["名称", "购买单价", "提前期"])
    for name, item in (params.get("materials") or {}).items(): ws.append([name, _money(item.get("price_wan")), f"{item.get('lead_quarters', 0)}季"])
    ws.append(["八.其它说明"])
    ws.append([f"1. 紧急采购付款即到货；原料价格为正常价格的 {params.get('emergency_material_price_multiplier', 0)} 倍，成品价格为直接成本的 {params.get('emergency_product_price_multiplier', 0)} 倍。"])
    ws.append(["2. 订单冲突优先级：" + " → ".join(str(value) for value in (params.get("selection_priority") or []))])
    ws.append(["3. 破产标准：现金断流或所有者权益为负；破产由环境结算判定。"])
    ws.append([f"4. 第一年度{'有' if params.get('first_year_has_orders') else '无'}订单；全赛程订单可提前查看，通常在释放前 1 个季度进入申领窗口。"])
    ws.append(["5. 订单可以提前交付，不可晚于交期；逾期由环境收回并按规则计罚。"])
    ws.append(["6. 舍入规则：" + json.dumps(params.get("rounding") or {}, ensure_ascii=False)])
    ws.append(["7. 库存折价、资产处置损失、紧急采购和订单违约均进入财务报表。"])
    ws.append(["8. 排行榜记分标准："])
    ws.append([params.get("score_formula") or "score = owner_equity * (1 + development_potential / 100)"])
    ws.append(["企业综合发展潜力由市场、ISO、产品、自有厂房和已建成生产线分值汇总。"])
    ws.append(["九.重要参数"])
    ws.append(["违约金比例", params.get("default_penalty_rate"), "贷款额倍数", long_loan.get("max_total_multiple_prior_equity")])
    ws.append(["长贷利率", long_loan.get("annual_rate"), "短贷利率", short_loan.get("rate")])
    ws.append(["1、2期贴现率", discount.get("terms_1_2"), "3、4期贴现率", discount.get("terms_3_4")])
    ws.append(["初始现金", _money(params.get("initial_cash_wan")), "管理费", _money(params.get("management_fee_per_quarter_wan"))])
    ws.append(["所得税率", params.get("tax_rate"), "最小得单广告额", _money(params.get("minimum_order_advertising_wan"))])
    ws.append(["原料紧急采购倍数", params.get("emergency_material_price_multiplier"), "产品紧急采购倍数", params.get("emergency_product_price_multiplier")])
    ws.append(["最大长贷年限", long_loan.get("max_years"), "最大厂房数量", asset_limits.get("max_factories_total", params.get("max_factory_count"))])
    for row in ws.iter_rows():
        if row and row[0].value and (str(row[0].value).startswith(tuple("一二三四五六七八九")) or row[0].value == "重要经营规则"):
            row[0].font = Font(bold=True, size=13)
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"
    for col in range(1, 13): ws.column_dimensions[get_column_letter(col)].width = 19
    _metadata_sheet(workbook, {"format_version": COMPETITION_XLSX_VERSION, "match_id": rules.get("match_id"), "rule_pack_id": rules.get("rule_pack_id"), "parent_rule_pack_id": rules.get("parent_rule_pack_id"), "rules_json": dict(rules), "provenance": "simulated"})
    workbook.save(path)


def _write_results(path: Path, arena: FullCompetitionArena) -> None:
    results = arena.final_results(); workbook = Workbook(); ws = workbook.active; ws.title = "最终排名"
    _header(ws, 1, ["排名", "企业", "所有者权益", "发展潜力", "总分", "是否破产", "破产时间"], 1)
    bankruptcy = {row["team_id"]: row for row in results.get("bankruptcies", [])}
    for row_index, row in enumerate(results.get("ranking", []), 2):
        broken = bankruptcy.get(row.get("team_id"), {})
        values = [row.get("rank"), row.get("team_id"), row.get("owner_equity_wan"), row.get("development_potential"), row.get("rounded_score", row.get("score")), "是" if broken else "否", broken.get("bankruptcy_period")]
        for col, value in enumerate(values, 1): ws.cell(row_index, col, value)
    _finish_sheet(ws, {1: 10, 2: 20, 3: 16, 4: 16, 5: 14, 6: 12, 7: 16})

    ws = workbook.create_sheet("破产信息")
    _header(ws, 1, ["企业", "破产时间", "终局现金", "终局权益", "破产原因"], 1)
    for row_index, state in enumerate(sorted((state for state in arena.states.values() if state.bankrupt), key=lambda item: (item.bankruptcy_period or "", item.team_id)), 2):
        values = [state.team_id, state.bankruptcy_period, state.cash_wan, state.owner_equity_wan, "、".join(state.bankruptcy_reasons)]
        for col, value in enumerate(values, 1): ws.cell(row_index, col, value)
    _finish_sheet(ws, {1: 20, 2: 16, 3: 14, 4: 14, 5: 52})

    ws = workbook.create_sheet("全部企业终局")
    _header(ws, 1, ["企业", "状态", "现金", "所有者权益", "获单数", "交付数", "违约数", "厂房数", "产线数", "破产时间"], 1)
    for row_index, state in enumerate(sorted(arena.states.values(), key=lambda item: item.team_id), 2):
        values = [
            state.team_id, "破产" if state.bankrupt else "经营结束", state.cash_wan, state.owner_equity_wan,
            len(state.assigned_orders), len(state.delivered_orders), len(state.defaulted_orders),
            len(state.factories), len(state.production_lines), state.bankruptcy_period,
        ]
        for col, value in enumerate(values, 1): ws.cell(row_index, col, value)
    _finish_sheet(ws, {1: 20, 2: 14, 3: 14, 4: 16, 5: 12, 6: 12, 7: 12, 8: 12, 9: 12, 10: 16})
    _metadata_sheet(workbook, {"format_version": COMPETITION_XLSX_VERSION, "match_id": results.get("match_id"), "provenance": "simulated"})
    workbook.save(path)


def _annual_advertising(arena: FullCompetitionArena, year: int) -> dict[tuple[str, str, str], float]:
    values: dict[tuple[str, str, str], float] = {}
    for team_id, state in arena.states.items():
        for event in state.journal:
            if event.get("event_type") != "advertising_expense":
                continue
            period = str(event.get("period") or "")
            if not period.startswith(f"Y{year}Q"):
                continue
            market, _, product = str(event.get("advertising_key") or "本地:P1").partition(":")
            key = (team_id, market or "本地", product or "P1")
            values[key] = values.get(key, 0.0) + abs(float(event.get("cash_effect_wan") or 0))
    return values


def _write_annual_public(path: Path, year: int, arena: FullCompetitionArena, rules: Mapping[str, Any]) -> None:
    params = rules.get("parameters", rules)
    markets = list((params.get("markets") or {}).keys())
    products = list((params.get("products") or {}).keys())
    team_ids = sorted(arena.states)
    advertising = _annual_advertising(arena, year)
    initial_cash = float(params.get("initial_cash_wan", 0))

    workbook = Workbook()
    ws = workbook.active
    ws.title = f"第{year}年广告投放"
    row = 1
    for team_id in team_ids:
        ws.cell(row, 2, f"{team_id}广告投放情况")
        _header(ws, row + 1, ["产品", *markets], 2)
        for offset, product in enumerate(products, row + 2):
            ws.cell(offset, 2, product)
            for market_col, market in enumerate(markets, 3):
                ws.cell(offset, market_col, advertising.get((team_id, market, product), 0.0))
        row += len(products) + 2
    _finish_sheet(ws, {2: 18, **{index: 13 for index in range(3, 3 + len(markets))}})

    ws = workbook.create_sheet(f"第{year}年广告投放(格式二)")
    row = 1
    for market in markets:
        _header(ws, row, ["产品用户", *team_ids], 1)
        ws.cell(row + 1, 1, f"{market}广告投放情况")
        for offset, product in enumerate(products, row + 2):
            ws.cell(offset, 1, product)
            for team_col, team_id in enumerate(team_ids, 2):
                ws.cell(offset, team_col, advertising.get((team_id, market, product), 0.0))
        row += len(products) + 3
    _finish_sheet(ws, {1: 18, **{index: 14 for index in range(2, 2 + len(team_ids))}})

    ws = workbook.create_sheet(f"第{year}年三张报表")
    _header(ws, 2, ["用户名", *team_ids], 2)
    report_values: dict[str, tuple[dict[str, float], dict[str, float], dict[str, float]]] = {}
    for team_id, state in arena.states.items():
        report = next((item for item in state.reports if int(item.get("year", -1)) == year), {})
        report_values[team_id] = _report_values(report, initial_cash)
    row = 3
    for section, metrics in ((0, EXPENSE_METRICS), (1, INCOME_METRICS), (2, BALANCE_METRICS)):
        for metric in metrics:
            ws.cell(row, 2, metric)
            for team_col, team_id in enumerate(team_ids, 3):
                ws.cell(row, team_col, report_values[team_id][section].get(metric, 0.0))
            row += 1
        row += 1
    _finish_sheet(ws, {2: 24, **{index: 14 for index in range(3, 3 + len(team_ids))}})

    ws = workbook.create_sheet("生产线信息")
    _header(ws, 2, ["所属用户", "名称", "厂房", "产品", "状态", "累计折旧", "开产时间", "转产时间", "剩余时间", "建成时间", "开建时间"], 2)
    row = 3
    for team_id, state in sorted(arena.states.items()):
        for line in state.production_lines:
            values = [team_id, line.get("line_type"), line.get("factory_id") or "-", line.get("product_id") or "-", line.get("status"), _money(line.get("accumulated_depreciation_wan")), "-", "-", "0季", _period_cn(line.get("completed_period")), _period_cn(line.get("ordered_period"))]
            for col, value in enumerate(values, 2):
                ws.cell(row, col, value)
            row += 1
    _finish_sheet(ws, {index: 16 for index in range(2, 13)})

    ws = workbook.create_sheet(f"第{year}年市场老大")
    _header(ws, 1, ["市场", "市场老大", "广告总额"], 2)
    for row, market in enumerate(markets, 2):
        totals = {
            team_id: sum(advertising.get((team_id, market, product), 0.0) for product in products)
            for team_id in team_ids
        }
        leader = min(team_ids, key=lambda team_id: (-totals[team_id], team_id)) if team_ids and max(totals.values(), default=0) > 0 else None
        ws.cell(row, 2, market); ws.cell(row, 3, leader); ws.cell(row, 4, totals.get(leader, 0.0) if leader else 0.0)
    _finish_sheet(ws, {2: 16, 3: 22, 4: 14})
    _metadata_sheet(workbook, {"format_version": COMPETITION_XLSX_VERSION, "match_id": rules.get("match_id"), "year": year, "role": "annual_public_global_export", "provenance": "simulated"})
    workbook.save(path)


def export_competition_xlsx(output_dir: Path, *, rules: Mapping[str, Any], orders: Sequence[Mapping[str, Any]], arena: FullCompetitionArena) -> dict[str, Any]:
    """Export one complete simulation as a competition-style XLSX bundle."""

    output_dir = output_dir.resolve(); enterprise_dir = output_dir / "企业数据"; enterprise_dir.mkdir(parents=True, exist_ok=True)
    match_id = str(rules.get("match_id")); team_files = []
    for team_id, state in sorted(arena.states.items()):
        path = enterprise_dir / f"{team_id}.xlsx"; _write_enterprise_workbook(path, state, rules); team_files.append(path.relative_to(output_dir).as_posix())
    merged_orders = _merge_global_orders(orders, arena)
    rules_name = f"{match_id}比赛规则.xlsx"; orders_name = f"{match_id}订单详情.xlsx"; results_name = f"{match_id}最终排名和破产信息.xlsx"
    _write_rules(output_dir / rules_name, rules); _write_global_orders(output_dir / orders_name, merged_orders, rules); _write_results(output_dir / results_name, arena)
    annual_public_files = []
    for year in range(1, 7):
        filename = f"{year}.xlsx"
        _write_annual_public(output_dir / filename, year, arena, rules)
        annual_public_files.append(filename)
    manifest = {"format_version": COMPETITION_XLSX_VERSION, "match_id": match_id, "rule_pack_id": rules.get("rule_pack_id"), "parent_rule_pack_id": rules.get("parent_rule_pack_id"), "source_match_id": (rules.get("generation") or {}).get("source_match_id"), "provenance": "simulated", "training_eligible": False, "enterprise_sheet_names": list(ENTERPRISE_SHEETS), "annual_public_sheet_roles": ["年度广告投放", "年度广告投放格式二", "年度三张报表", "生产线信息", "年度市场老大"], "team_count": len(team_files), "order_count": len(merged_orders), "files": {"enterprise": team_files, "annual_public": annual_public_files, "rules": rules_name, "orders": orders_name, "results": results_name}}
    _json(output_dir / "manifest.json", manifest)
    return manifest


def build_competition_xlsx_archive(*, rules: Mapping[str, Any], orders: Sequence[Mapping[str, Any]], arena: FullCompetitionArena) -> bytes:
    """Return a ZIP whose layout mirrors a complete Chinese competition export."""

    match_id = str(rules.get("match_id") or "SIM_MATCH")
    with tempfile.TemporaryDirectory(prefix="stratpilot-xlsx-") as directory:
        bundle = Path(directory) / "generated"
        manifest = export_competition_xlsx(bundle, rules=rules, orders=orders, arena=arena)
        archive_manifest = dict(manifest)
        archive_manifest["archive_format_version"] = COMPETITION_ARCHIVE_VERSION
        archive_manifest["files"] = {
            "enterprise": [f"{match_id}/{Path(name).name}" for name in manifest["files"]["enterprise"]],
            "annual_public": [f"{match_id}/{Path(name).name}" for name in manifest["files"]["annual_public"]],
            "rules": manifest["files"]["rules"],
            "orders": manifest["files"]["orders"],
            "results": manifest["files"]["results"],
        }
        readme = (
            f"{match_id} 模拟比赛资料包\n\n"
            f"{match_id}/：第1年至第6年公共表，以及每家企业的完整工作簿。\n"
            f"{manifest['files']['rules']}：本场实际执行规则。\n"
            f"{manifest['files']['orders']}：全局订单及终局归属、状态。\n"
            f"{manifest['files']['results']}：最终排名、权益、发展潜力和破产时间。\n"
            "所有文件均由本场模拟器生成，provenance=simulated，不代表历史赛事官方导出。\n"
        )
        output = BytesIO()
        with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=6) as archive:
            for relative in manifest["files"]["enterprise"]:
                archive.write(bundle / relative, f"{match_id}/{Path(relative).name}")
            for relative in manifest["files"]["annual_public"]:
                archive.write(bundle / relative, f"{match_id}/{Path(relative).name}")
            for key in ("rules", "orders", "results"):
                relative = manifest["files"][key]
                archive.write(bundle / relative, Path(relative).name)
            archive.writestr("manifest.json", json.dumps(archive_manifest, ensure_ascii=False, indent=2) + "\n")
            archive.writestr("导出说明.txt", readme)
        return output.getvalue()


def _meta(workbook) -> dict[str, Any]:
    if "_GOAI_META" not in workbook.sheetnames: return {}
    output = {}
    for key, value in workbook["_GOAI_META"].iter_rows(min_row=2, values_only=True):
        if key is None: continue
        try: output[str(key)] = json.loads(value) if isinstance(value, str) else value
        except json.JSONDecodeError: output[str(key)] = value
    return output


class SimulatedCompetitionXlsxImporter:
    """Read a generated XLSX bundle back into the normalized JSONL shape."""

    def __init__(self, bundle_dir: Path, output_dir: Path) -> None:
        self.bundle_dir = bundle_dir.resolve(); self.output_dir = output_dir.resolve()

    def import_bundle(self) -> dict[str, Any]:
        manifest = json.loads((self.bundle_dir / "manifest.json").read_text(encoding="utf-8"))
        if manifest.get("provenance") != "simulated": raise ValueError("only explicitly simulated XLSX bundles are accepted")
        match_id = str(manifest["match_id"]); teams = []; events = []; team_orders = []; reports = []
        for relative in manifest["files"]["enterprise"]:
            path = self.bundle_dir / relative; workbook = load_workbook(path, data_only=True)
            missing = [name for name in ENTERPRISE_SHEETS if name not in workbook.sheetnames]
            if missing: raise ValueError(f"{path.name} missing sheets: {missing}")
            meta = _meta(workbook); team_id = str(meta.get("team_id") or path.stem)
            teams.append(self._team(match_id, team_id, path, workbook)); events.extend(self._events(match_id, team_id, path, workbook)); team_orders.extend(self._orders(match_id, team_id, path, workbook)); reports.extend(self._reports(match_id, team_id, path, workbook))
        global_orders = self._global_orders(match_id, self.bundle_dir / manifest["files"]["orders"])
        results = self._results(match_id, self.bundle_dir / manifest["files"]["results"])
        rules_workbook = load_workbook(self.bundle_dir / manifest["files"]["rules"], data_only=True); rules = _meta(rules_workbook).get("rules_json")
        if not isinstance(rules, dict): raise ValueError("rule workbook does not contain simulated rule metadata")
        self.output_dir.mkdir(parents=True, exist_ok=True)
        normalized_manifest = {**manifest, "format_version": "goai_simulated_xlsx_import_v1.0", "source_bundle": str(self.bundle_dir), "counts": {"teams": len(teams), "events": len(events), "global_orders": len(global_orders), "team_orders": len(team_orders), "reports": len(reports)}, "provenance": "simulated"}
        _json(self.output_dir / "manifest.json", normalized_manifest); _json(self.output_dir / "rules.json", rules); _jsonl(self.output_dir / "teams.jsonl", teams); _jsonl(self.output_dir / "events.jsonl", events); _jsonl(self.output_dir / "global_orders.jsonl", global_orders); _jsonl(self.output_dir / "team_orders.jsonl", team_orders); _jsonl(self.output_dir / "reports.jsonl", reports); _json(self.output_dir / "results.json", results)
        return normalized_manifest

    def _team(self, match_id: str, team_id: str, path: Path, workbook) -> dict[str, Any]:
        ws = workbook["企业信息"]
        return {"match_id": match_id, "team_id": team_id, "company_name": clean_text(ws["C6"].value), "school_name": clean_text(ws["E6"].value), "company_status": clean_text(ws["E4"].value), "export_cash_wan": parse_money_wan(ws["C4"].value), "capital_injection_wan": parse_money_wan(ws["C5"].value), "export_period": "Y6Q1", "data_status": "simulated_complete_export", "source_path": path.relative_to(self.bundle_dir).as_posix(), "source_sheet": "企业信息", "provenance": "simulated"}

    def _events(self, match_id: str, team_id: str, path: Path, workbook) -> list[dict[str, Any]]:
        ws = workbook["现金流量表"]; output = []
        for row in range(4, ws.max_row + 1):
            raw_action = clean_text(ws.cell(row, 3).value)
            if not raw_action: continue
            year, quarter = parse_period(ws.cell(row, 6).value); amount = as_number(ws.cell(row, 4).value); balance = as_number(ws.cell(row, 5).value)
            note = clean_text(ws.cell(row, 7).value)
            encoded_event = note.partition(" | ")[0] if note and " | " in note else None
            action = "capital_injection" if raw_action == "Pay_Capital" else (encoded_event or ACTION_EVENTS.get(raw_action, raw_action))
            output.append({"match_id": match_id, "event_id": f"{match_id}:{team_id}:{row - 3}", "team_id": team_id, "transaction_id": as_int(ws.cell(row, 2).value), "sequence_in_source": row - 3, "period": f"Y{year}Q{quarter}" if year and quarter else None, "year": year, "quarter": quarter, "action_raw": raw_action, "action": action, "amount_wan": amount, "balance_wan": balance, "note": note, "included_in_match": bool(year in range(1, 6) and quarter in range(1, 5)), "source_path": path.relative_to(self.bundle_dir).as_posix(), "source_sheet": "现金流量表", "source_row": row, "provenance": "simulated"})
        return output

    def _orders(self, match_id: str, team_id: str, path: Path, workbook) -> list[dict[str, Any]]:
        ws = workbook["订单信息"]; output = []
        for row in range(4, ws.max_row + 1):
            order_id = clean_text(ws.cell(row, 2).value)
            if not order_id: continue
            year, _ = parse_period(ws.cell(row, 8).value); delivered_year, delivered_quarter = parse_period(ws.cell(row, 12).value)
            output.append({"match_id": match_id, "order_id": order_id, "year": year, "market": clean_text(ws.cell(row, 3).value), "product": clean_text(ws.cell(row, 4).value), "quantity": as_number(ws.cell(row, 5).value), "total_price_wan": parse_money_wan(ws.cell(row, 6).value), "status": clean_text(ws.cell(row, 7).value), "delivery_term_quarters": as_number(ws.cell(row, 9).value), "receivable_term_quarters": as_number(ws.cell(row, 10).value), "iso": clean_text(ws.cell(row, 11).value) or "-", "owner_team_id": team_id, "delivered_period": f"Y{delivered_year}Q{delivered_quarter}" if delivered_year and delivered_quarter else None, "coverage_scope": "enterprise_allocated_order", "source_path": path.relative_to(self.bundle_dir).as_posix(), "source_sheet": "订单信息", "source_row": row, "provenance": "simulated"})
        return output

    def _reports(self, match_id: str, team_id: str, path: Path, workbook) -> list[dict[str, Any]]:
        ws = workbook["三张报表"]; output = []
        for header_row, start, stop, statement in ((2, 3, 14, "comprehensive_expense"), (15, 16, 27, "income_statement"), (28, 30, 51, "balance_sheet")):
            for col in range(3, min(ws.max_column, 8) + 1):
                period = clean_text(ws.cell(header_row, col).value); year, _ = parse_period(period)
                if year is None: continue
                for row in range(start, min(stop, ws.max_row + 1)):
                    metric = clean_text(ws.cell(row, 2).value); value = as_number(ws.cell(row, col).value)
                    if metric and value is not None: output.append({"match_id": match_id, "team_id": team_id, "year": year, "period_label": period, "statement": statement, "metric": metric, "value_wan": value, "report_variant": "system", "source_path": path.relative_to(self.bundle_dir).as_posix(), "source_sheet": "三张报表", "source_cell": f"{get_column_letter(col)}{row}", "provenance": "simulated"})
        return output

    def _global_orders(self, match_id: str, path: Path) -> list[dict[str, Any]]:
        ws = load_workbook(path, data_only=True)["订单详情"]; output = []
        for row in range(2, ws.max_row + 1):
            order_id = clean_text(ws.cell(row, 1).value)
            if not order_id: continue
            output.append({"match_id": match_id, "order_id": order_id, "order_type": clean_text(ws.cell(row, 2).value), "year": as_int(ws.cell(row, 3).value), "market": clean_text(ws.cell(row, 4).value), "product": clean_text(ws.cell(row, 5).value), "quantity": as_number(ws.cell(row, 6).value), "total_price_wan": as_number(ws.cell(row, 7).value), "delivery_term_quarters": as_number(ws.cell(row, 8).value), "receivable_term_quarters": as_number(ws.cell(row, 9).value), "iso": clean_text(ws.cell(row, 10).value) or "-", "owner_team_id": clean_text(ws.cell(row, 11).value), "status": clean_text(ws.cell(row, 12).value), "coverage_scope": "simulated_global_order_pool", "source_path": path.relative_to(self.bundle_dir).as_posix(), "source_sheet": "订单详情", "source_row": row, "provenance": "simulated"})
        return output

    def _results(self, match_id: str, path: Path) -> dict[str, Any]:
        ws = load_workbook(path, data_only=True)["最终排名"]; ranking = []; bankruptcies = []
        for row in range(2, ws.max_row + 1):
            team_id = clean_text(ws.cell(row, 2).value)
            if not team_id: continue
            item = {"rank": as_int(ws.cell(row, 1).value), "team_id": team_id, "owner_equity_wan": as_number(ws.cell(row, 3).value), "development_potential": as_number(ws.cell(row, 4).value), "rounded_score": as_number(ws.cell(row, 5).value)}; ranking.append(item)
            if clean_text(ws.cell(row, 6).value) == "是": bankruptcies.append({"team_id": team_id, "bankruptcy_period": clean_text(ws.cell(row, 7).value)})
        return {"match_id": match_id, "ranking": ranking, "bankruptcies": bankruptcies, "provenance": "simulated", "source_path": path.relative_to(self.bundle_dir).as_posix()}
