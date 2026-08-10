from __future__ import annotations

import contextlib
import io
import json
import math
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

import xlrd
from docx import Document
from openpyxl import load_workbook

from .common import as_int, as_number, cell_ref, clean_text, parse_duration, parse_money_wan, parse_period, sha256_file, stable_record_id
from .rulepack import ACTION_CONTROL_TYPES, ACTION_DEFINITIONS, canonicalize_action, parse_action_parameters


FORMAT_VERSION = "goai_multimatch_v2.0"
OPERATING_YEARS = range(1, 6)
OPERATING_QUARTERS = range(1, 5)
OLD_MATCH_IDS = ("AB", "AG", "CA", "CB", "CD", "CE", "EA", "EB", "EC", "EF", "OP", "ZY", "ZZ")
OUTPUT_FILES = (
    "manifest.json",
    "rules.json",
    "teams.jsonl",
    "events.jsonl",
    "global_orders.jsonl",
    "annual_public.jsonl",
    "reports.jsonl",
    "final_states.jsonl",
    "quarter_states.jsonl",
    "results.json",
    "raw_cells.jsonl",
    "quality.json",
)

XA_OFFICIAL_RANKING = (
    ("XA07", 3791), ("XA13", 2551), ("XA06", 2506), ("XA08", 2393), ("XA22", 2353), ("XA10", 2144),
    ("XA15", 1840), ("XA19", 1764), ("XA02", 1451), ("XA01", 1295), ("XA03", 1271), ("XA12", 1198),
    ("XA25", 1164), ("XA18", 1076), ("XA27", 736), ("XA21", 405), ("XA11", 242), ("XA26", 69),
)

CHINESE_ACTION_ALIASES = {
    "初始化资本金": "capital_injection",
    "支付行政管理费": "administrative_fee",
    "开始下一批生产": "production",
    "支付设备维修费用": "maintenance",
    "在建生产线": "production_line_investment",
    "更新原材料": "material_receipt_payment",
    "短贷": "short_loan_borrow",
    "短贷利息": "short_loan_interest_payment",
    "归还短贷": "short_loan_principal_payment",
    "产品研发": "product_development",
    "新建生产线": "production_line_order",
    "应收款更新": "receivable_maturity",
    "贴现": "receivable_discount",
    "厂房租金": "factory_rent_renewal",
    "ISO投资": "iso_development",
    "广告投放": "advertising",
    "市场开拓": "market_development",
    "长贷利息": "long_loan_interest_payment",
    "归还长贷": "long_loan_principal_payment",
    "长贷": "long_loan_borrow",
    "转产": "production_line_conversion",
    "厂房租用": "factory_rent",
    "订单交货": "order_delivery",
    "厂房购买": "factory_purchase",
    "出售库存": "inventory_sale",
    "缴纳所得税": "tax_payment",
    "紧急采购": "emergency_purchase",
    "厂房处理": "factory_disposal",
    "出售生产线": "production_line_sale",
    "竞单会标书费": "order_bid_fee",
    "厂房贴现": "factory_discount",
    "订单违约支付罚金": "penalty_payment",
    "间谍": "spy_information_purchase",
}

EXTRA_ACTION_METADATA = {
    "long_loan_interest_payment": ("finance", "committed_settlement"),
    "long_loan_principal_payment": ("finance", "committed_settlement"),
    "inventory_sale": ("sales", "conditional_decision"),
    "emergency_purchase": ("procurement", "exception_intervention"),
    "factory_disposal": ("capacity", "conditional_decision"),
    "production_line_sale": ("capacity", "conditional_decision"),
    "order_bid_fee": ("sales", "committed_settlement"),
    "factory_discount": ("finance", "conditional_decision"),
    "spy_information_purchase": ("information", "direct_decision"),
    "factory_sale": ("capacity", "conditional_decision"),
    "inventory_material_sale": ("sales", "conditional_decision"),
    "equity_financing": ("finance", "exception_intervention"),
}

ENGLISH_EXTRA_ALIASES = {
    "Workshop_Discount": "factory_discount",
    "Sell_Workshop": "factory_sale",
    "Spy": "spy_information_purchase",
    "Sell_Inventory_Material": "inventory_material_sale",
    "Sell_ProductLine": "production_line_sale",
    "Financing_Add": "equity_financing",
}


def _open_xls(path: Path) -> xlrd.book.Book:
    # A few source files have harmless OLE allocation warnings printed by xlrd.
    with contextlib.redirect_stdout(io.StringIO()), contextlib.redirect_stderr(io.StringIO()):
        return xlrd.open_workbook(str(path))


def _json_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _period_label(year: int | None, quarter: int | None) -> str | None:
    if year is None:
        return None
    return f"Y{year}Q{quarter}" if quarter is not None else f"Y{year}"


def _value_right_of(sheet: xlrd.sheet.Sheet, label: str) -> Any:
    for row in range(sheet.nrows):
        for col in range(sheet.ncols - 1):
            if clean_text(sheet.cell_value(row, col)) == label:
                return sheet.cell_value(row, col + 1)
    return None


def _write_json(path: Path, value: Any) -> None:
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def _write_jsonl(path: Path, rows: Iterable[dict[str, Any]]) -> None:
    text = "".join(json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n" for row in rows)
    path.write_text(text, encoding="utf-8")


class MultiMatchDatasetBuilder:
    def __init__(self, project_root: Path, output_root: Path) -> None:
        self.project_root = project_root.resolve()
        self.output_root = output_root.resolve()
        self.raw_root = self.project_root / "data/original/VPD-OE_Agent_20260709/VPD-OE_Agent/data/raw"
        self.xa_bundle_root = self.project_root / "data/original/lx_26140727378文件夹对应的省赛一个赛场所有资料"
        self.xa_raw_root = self.xa_bundle_root / "lx_26140727378"
        self.old_candidate_rules = self.project_root / "data/original/VPD-OE_Agent_20260709/VPD-OE_Agent/data/rules/competition_rules.json"

    def build(self) -> dict[str, Any]:
        self.output_root.mkdir(parents=True, exist_ok=True)
        matches_root = self.output_root / "matches"
        matches_root.mkdir(exist_ok=True)
        summaries = []
        for match_id in (*OLD_MATCH_IDS, "LX_XA"):
            summaries.append(self._build_match(match_id, matches_root / match_id))
        catalog = {
            "format_version": FORMAT_VERSION,
            "dataset_id": "goai_chinese_enterprise_competitions_v2",
            "match_count": len(summaries),
            "matches": summaries,
        }
        _write_json(self.output_root / "catalog.json", catalog)
        self._write_readme(catalog)
        return catalog

    def _source_root(self, match_id: str) -> Path:
        return self.xa_raw_root if match_id == "LX_XA" else self.raw_root / match_id

    def _extra_sources(self, match_id: str) -> list[Path]:
        if match_id != "LX_XA":
            return []
        return [
            self.xa_bundle_root / "lx_26140727378文件夹对应的省赛规则.xlsx",
            self.xa_bundle_root / "lx_26140727378文件夹对应的省赛订单详情.xlsx",
            self.xa_bundle_root / "lx_26140727378文件夹对应的最终排名和破产组破产时间信息.docx",
        ]

    def _build_match(self, match_id: str, out: Path) -> dict[str, Any]:
        out.mkdir(parents=True, exist_ok=True)
        source_root = self._source_root(match_id)
        xls_files = sorted(source_root.glob("*.xls"))
        enterprise_files, public_files = self._classify_xls_files(xls_files)
        extra_sources = self._extra_sources(match_id)
        sources = xls_files + [path for path in extra_sources if path.is_file()]
        manifest_sources = [self._source_record(path) for path in sources]

        raw_cells: list[dict[str, Any]] = []
        teams: list[dict[str, Any]] = []
        events: list[dict[str, Any]] = []
        orders: list[dict[str, Any]] = []
        reports: list[dict[str, Any]] = []
        annual_public: list[dict[str, Any]] = []
        assets: dict[str, dict[str, Any]] = {}

        for path in xls_files:
            workbook = _open_xls(path)
            raw_cells.extend(self._raw_xls_cells(match_id, path, workbook))
        for path in extra_sources:
            if path.suffix.lower() == ".xlsx":
                raw_cells.extend(self._raw_xlsx_cells(match_id, path))

        for path in enterprise_files:
            workbook = _open_xls(path)
            team_id = path.stem.upper()
            teams.append(self._parse_team(match_id, team_id, path, workbook))
            events.extend(self._parse_events(match_id, team_id, path, workbook))
            orders.extend(self._parse_team_orders(match_id, team_id, path, workbook))
            reports.extend(self._parse_team_reports(match_id, team_id, path, workbook))
            reports.extend(self._parse_team_advertising(match_id, team_id, path, workbook))
            assets[team_id] = self._parse_score_assets(team_id, workbook)

        annual_public.extend(self._parse_public_files(match_id, public_files))
        known_team_ids = self._team_ids_from_public(public_files)
        existing = {team["team_id"] for team in teams}
        for team_id in sorted(known_team_ids - existing):
            teams.append(self._missing_team(match_id, team_id))

        if match_id == "LX_XA":
            global_orders = self._parse_xa_global_orders(extra_sources[1])
            rules = self._parse_xa_rules(extra_sources[0])
            bankruptcies = self._parse_xa_bankruptcies(extra_sources[2])
        else:
            global_orders = self._reconstruct_order_pool(match_id, orders)
            rules = self._build_old_rules(match_id, events, enterprise_files, global_orders)
            bankruptcies = {}

        quarter_states = self._build_quarter_states(match_id, teams, events, bankruptcies)
        final_states = self._build_final_states(match_id, teams, events, reports, assets, rules, bankruptcies)
        results = self._build_results(match_id, final_states, bankruptcies, rules)
        quality = self._quality(match_id, teams, events, orders, global_orders, reports, quarter_states, sources)

        manifest = {
            "format_version": FORMAT_VERSION,
            "match_id": match_id,
            "source_kind": "complete_provincial_bundle" if match_id == "LX_XA" else "historical_enterprise_and_public_exports",
            "operating_period": {"start": "Y1Q1", "end": "Y5Q4", "export_boundary": "Y6Q1"},
            "schema_complete": True,
            "observed_complete": match_id == "LX_XA",
            "truth_policy": "Never promote inferred, simulated or missing values to observed facts.",
            "provenance_values": ["observed", "derived", "inferred", "simulated", "missing"],
            "files": list(OUTPUT_FILES),
            "sources": manifest_sources,
        }
        _write_json(out / "manifest.json", manifest)
        _write_json(out / "rules.json", rules)
        _write_jsonl(out / "teams.jsonl", sorted(teams, key=lambda row: row["team_id"]))
        _write_jsonl(out / "events.jsonl", sorted(events, key=lambda row: (row["team_id"], row["sequence_in_source"])))
        _write_jsonl(out / "global_orders.jsonl", sorted(global_orders, key=lambda row: row["order_id"]))
        _write_jsonl(out / "annual_public.jsonl", annual_public)
        _write_jsonl(out / "reports.jsonl", reports)
        _write_jsonl(out / "final_states.jsonl", sorted(final_states, key=lambda row: row["team_id"]))
        _write_jsonl(out / "quarter_states.jsonl", sorted(quarter_states, key=lambda row: (row["team_id"], row["period_index"])))
        _write_json(out / "results.json", results)
        _write_jsonl(out / "raw_cells.jsonl", raw_cells)
        _write_json(out / "quality.json", quality)
        return {
            "match_id": match_id,
            "team_count": len(teams),
            "event_count": len(events),
            "operating_event_count": sum(row["included_in_match"] for row in events),
            "global_order_count": len(global_orders),
            "observed_order_count": sum(row.get("provenance") == "observed" for row in global_orders),
            "simulated_order_count": sum(row.get("provenance") == "simulated" for row in global_orders),
            "global_order_coverage": quality["coverage"]["global_orders"],
            "cash_continuity_passed": quality["checks"]["cash_continuity"]["passed"],
            "observed_complete": match_id == "LX_XA",
            "path": f"matches/{match_id}",
        }

    @staticmethod
    def _classify_xls_files(files: list[Path]) -> tuple[list[Path], list[Path]]:
        enterprise, public = [], []
        for path in files:
            workbook = _open_xls(path)
            if "现金流量表" in workbook.sheet_names():
                enterprise.append(path)
            else:
                public.append(path)
        return enterprise, public

    def _source_record(self, path: Path) -> dict[str, Any]:
        return {
            "source_id": "src_" + stable_record_id(path.as_posix(), sha256_file(path)),
            "path": path.relative_to(self.project_root).as_posix(),
            "size_bytes": path.stat().st_size,
            "sha256": sha256_file(path),
        }

    def _raw_xls_cells(self, match_id: str, path: Path, workbook: xlrd.book.Book) -> list[dict[str, Any]]:
        rows = []
        rel = path.relative_to(self.project_root).as_posix()
        for sheet in workbook.sheets():
            for row in range(sheet.nrows):
                for col in range(sheet.ncols):
                    value = sheet.cell_value(row, col)
                    if clean_text(value) is None:
                        continue
                    rows.append({
                        "match_id": match_id,
                        "source_path": rel,
                        "sheet": sheet.name,
                        "cell": cell_ref(row, col),
                        "row": row + 1,
                        "column": col + 1,
                        "value": _json_value(value),
                        "provenance": "observed",
                    })
        return rows

    def _raw_xlsx_cells(self, match_id: str, path: Path) -> list[dict[str, Any]]:
        rows = []
        workbook = load_workbook(path, data_only=False, read_only=False)
        rel = path.relative_to(self.project_root).as_posix()
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if clean_text(cell.value) is None:
                        continue
                    rows.append({
                        "match_id": match_id,
                        "source_path": rel,
                        "sheet": sheet.title,
                        "cell": cell.coordinate,
                        "row": cell.row,
                        "column": cell.column,
                        "value": _json_value(cell.value),
                        "provenance": "observed",
                    })
        return rows

    def _parse_team(self, match_id: str, team_id: str, path: Path, workbook: xlrd.book.Book) -> dict[str, Any]:
        sheet = workbook.sheet_by_name("企业信息")
        cash_sheet = workbook.sheet_by_name("现金流量表")
        system_year, system_quarter = parse_period(_value_right_of(sheet, "系统时间"))
        company_name = _value_right_of(sheet, "公司名称")
        school_name = _value_right_of(sheet, "所属学校")
        if school_name is None:
            school_name = _value_right_of(sheet, "学校")
        has_events = any(clean_text(cash_sheet.cell_value(row, 2)) for row in range(3, cash_sheet.nrows))
        return {
            "match_id": match_id,
            "team_id": team_id,
            "company_name": clean_text(company_name),
            "school_name": clean_text(school_name),
            "company_status": clean_text(_value_right_of(sheet, "公司状态")),
            "export_cash_wan": parse_money_wan(_value_right_of(sheet, "公司现金")),
            "capital_injection_wan": parse_money_wan(_value_right_of(sheet, "股东注资")),
            "export_period": _period_label(system_year, system_quarter),
            "data_status": "active_export" if has_events else "empty_export",
            "source_path": path.relative_to(self.project_root).as_posix(),
            "source_sheet": sheet.name,
            "provenance": "observed",
        }

    @staticmethod
    def _missing_team(match_id: str, team_id: str) -> dict[str, Any]:
        return {
            "match_id": match_id,
            "team_id": team_id,
            "company_name": None,
            "school_name": None,
            "company_status": None,
            "export_cash_wan": None,
            "capital_injection_wan": None,
            "export_period": None,
            "data_status": "public_only_missing_enterprise_export",
            "source_path": None,
            "source_sheet": None,
            "provenance": "missing",
        }

    def _parse_events(self, match_id: str, team_id: str, path: Path, workbook: xlrd.book.Book) -> list[dict[str, Any]]:
        sheet = workbook.sheet_by_name("现金流量表")
        rows = []
        rel = path.relative_to(self.project_root).as_posix()
        for row in range(3, sheet.nrows):
            raw_action = clean_text(sheet.cell_value(row, 2))
            if not raw_action:
                continue
            year, quarter = parse_period(sheet.cell_value(row, 5))
            note = clean_text(sheet.cell_value(row, 6))
            canonical = canonicalize_action(raw_action, "historical") or CHINESE_ACTION_ALIASES.get(raw_action) or ENGLISH_EXTRA_ALIASES.get(raw_action)
            if raw_action in CHINESE_ACTION_ALIASES:
                parameters, parse_status = {}, "partial"
            else:
                parameters, parse_status = parse_action_parameters(raw_action, note, "historical")
            standard_metadata = ACTION_DEFINITIONS.get(canonical or "", {})
            extra_metadata = EXTRA_ACTION_METADATA.get(canonical or "")
            included = year in OPERATING_YEARS and quarter in OPERATING_QUARTERS
            rows.append({
                "match_id": match_id,
                "event_id": f"{match_id}:{team_id}:{as_int(sheet.cell_value(row, 1)) or row + 1}",
                "team_id": team_id,
                "transaction_id": as_int(sheet.cell_value(row, 1)),
                "sequence_in_source": row - 2,
                "period": _period_label(year, quarter),
                "year": year,
                "quarter": quarter,
                "action_raw": raw_action,
                "action": canonical,
                "action_category": standard_metadata.get("category") or (extra_metadata[0] if extra_metadata else None),
                "control_type": ACTION_CONTROL_TYPES.get(canonical or "") or (extra_metadata[1] if extra_metadata else None),
                "amount_wan": parse_money_wan(sheet.cell_value(row, 3)),
                "balance_wan": parse_money_wan(sheet.cell_value(row, 4)),
                "note": note,
                "parameters": parameters,
                "parameter_parse_status": parse_status,
                "included_in_match": included,
                "exclusion_reason": None if included else "outside_Y1Q1_Y5Q4_operating_window",
                "source_path": rel,
                "source_sheet": sheet.name,
                "source_row": row + 1,
                "provenance": "observed",
            })
        return rows

    def _parse_team_orders(self, match_id: str, team_id: str, path: Path, workbook: xlrd.book.Book) -> list[dict[str, Any]]:
        sheet = workbook.sheet_by_name("订单信息")
        rows = []
        for row in range(3, sheet.nrows):
            order_id = clean_text(sheet.cell_value(row, 1))
            if not order_id:
                continue
            won_year, _ = parse_period(sheet.cell_value(row, 7))
            delivered_year, delivered_quarter = parse_period(sheet.cell_value(row, 11))
            delivery_term, _, _ = parse_duration(sheet.cell_value(row, 8))
            receivable_term, _, _ = parse_duration(sheet.cell_value(row, 9))
            rows.append({
                "match_id": match_id,
                "order_id": order_id,
                "order_type": None,
                "year": won_year,
                "market": clean_text(sheet.cell_value(row, 2)),
                "product": clean_text(sheet.cell_value(row, 3)),
                "quantity": as_number(sheet.cell_value(row, 4)),
                "total_price_wan": parse_money_wan(sheet.cell_value(row, 5)),
                "delivery_term_quarters": delivery_term,
                "receivable_term_quarters": receivable_term,
                "iso": clean_text(sheet.cell_value(row, 10)) or "-",
                "owner_team_id": team_id,
                "status": clean_text(sheet.cell_value(row, 6)),
                "delivered_period": _period_label(delivered_year, delivered_quarter),
                "coverage_scope": "enterprise_allocated_order",
                "source_path": path.relative_to(self.project_root).as_posix(),
                "source_sheet": sheet.name,
                "source_row": row + 1,
                "provenance": "observed",
            })
        return rows

    def _parse_team_reports(self, match_id: str, team_id: str, path: Path, workbook: xlrd.book.Book) -> list[dict[str, Any]]:
        sheet_name = "三张报表" if "三张报表" in workbook.sheet_names() else "企业财务报表"
        sheet = workbook.sheet_by_name(sheet_name)
        rows = []
        sections = [(1, 2, 13, "comprehensive_expense"), (14, 15, 26, "income_statement"), (27, 28, 50, "balance_sheet")]
        is_paired = sheet_name == "企业财务报表"
        for header_row, start, stop, statement in sections:
            for col in range(2, sheet.ncols):
                period = clean_text(sheet.cell_value(header_row, col))
                if not period:
                    continue
                year, _ = parse_period(period)
                variant = clean_text(sheet.cell_value(header_row + 1, col)) if is_paired else "system"
                metric_start = start + 1 if is_paired and statement == "balance_sheet" else start
                for row in range(metric_start, min(stop, sheet.nrows)):
                    metric = clean_text(sheet.cell_value(row, 1))
                    if not metric or metric in {"年度", "类型"}:
                        continue
                    value = parse_money_wan(sheet.cell_value(row, col))
                    if value is None:
                        continue
                    rows.append({
                        "match_id": match_id,
                        "team_id": team_id,
                        "year": year,
                        "period_label": period,
                        "statement": statement,
                        "metric": metric,
                        "value_wan": value,
                        "report_variant": variant or "system",
                        "source_path": path.relative_to(self.project_root).as_posix(),
                        "source_sheet": sheet.name,
                        "source_cell": cell_ref(row, col),
                        "provenance": "observed",
                    })
        return rows

    def _parse_team_advertising(self, match_id: str, team_id: str, path: Path, workbook: xlrd.book.Book) -> list[dict[str, Any]]:
        if "广告投放" not in workbook.sheet_names():
            return []
        sheet = workbook.sheet_by_name("广告投放")
        rows = []
        row = 0
        while row < sheet.nrows:
            title = clean_text(sheet.cell_value(row, 1)) if sheet.ncols > 1 else None
            match = re.search(r"第(\d+)年广告投放情况", title or "")
            if not match:
                row += 1
                continue
            year = int(match.group(1))
            markets = [clean_text(sheet.cell_value(row + 1, col)) for col in range(2, sheet.ncols)]
            data_row = row + 2
            while data_row < sheet.nrows:
                product = clean_text(sheet.cell_value(data_row, 1))
                if not product or not product.startswith("P"):
                    break
                for col, market_name in enumerate(markets, start=2):
                    if market_name:
                        rows.append({
                            "match_id": match_id,
                            "team_id": team_id,
                            "year": year,
                            "period_label": f"Y{year}",
                            "statement": "advertising",
                            "metric": f"{market_name}:{product}",
                            "value_wan": parse_money_wan(sheet.cell_value(data_row, col)) or 0.0,
                            "report_variant": "enterprise_export",
                            "source_path": path.relative_to(self.project_root).as_posix(),
                            "source_sheet": sheet.name,
                            "source_cell": cell_ref(data_row, col),
                            "provenance": "observed",
                        })
                data_row += 1
            row = max(data_row, row + 1)
        return rows

    def _parse_public_files(self, match_id: str, files: list[Path]) -> list[dict[str, Any]]:
        records = []
        for path in files:
            workbook = _open_xls(path)
            year = as_int(path.stem) or as_int(re.search(r"(\d+)", path.stem).group(1) if re.search(r"(\d+)", path.stem) else None)
            for sheet in workbook.sheets():
                sheet_year_match = re.search(r"第(\d+)年", sheet.name)
                sheet_year = int(sheet_year_match.group(1)) if sheet_year_match else year
                section = self._public_section(sheet.name)
                if "广告投放(格式二)" in sheet.name:
                    records.extend(self._parse_public_advertising(match_id, path, sheet, sheet_year))
                elif "三张报表" in sheet.name:
                    records.extend(self._parse_public_combined_reports(match_id, path, sheet, sheet_year))
                elif any(name in sheet.name for name in ("综合费用表", "利润表", "资产负债表")):
                    records.extend(self._parse_public_separate_report(match_id, path, sheet, sheet_year, section))
                elif "市场老大" in sheet.name:
                    records.extend(self._parse_public_market_leader(match_id, path, sheet, sheet_year))
                elif sheet.name == "生产线信息":
                    records.extend(self._parse_public_rows(match_id, path, sheet, sheet_year, "production_lines"))
                else:
                    records.extend(self._parse_public_rows(match_id, path, sheet, sheet_year, section))
        return records

    @staticmethod
    def _public_section(sheet_name: str) -> str:
        if "广告" in sheet_name:
            return "advertising"
        if "综合费用" in sheet_name:
            return "comprehensive_expense"
        if "利润" in sheet_name:
            return "income_statement"
        if "资产负债" in sheet_name:
            return "balance_sheet"
        if "市场老大" in sheet_name:
            return "market_leader"
        return "raw_public_table"

    def _parse_public_advertising(self, match_id: str, path: Path, sheet: xlrd.sheet.Sheet, year: int | None) -> list[dict[str, Any]]:
        team_ids = [clean_text(sheet.cell_value(0, col)) for col in range(1, sheet.ncols)]
        market = None
        rows = []
        for row in range(1, sheet.nrows):
            label = clean_text(sheet.cell_value(row, 0))
            if label and label.endswith("广告投放情况"):
                market = label.removesuffix("广告投放情况")
                continue
            if not label or not label.startswith("P") or market is None:
                continue
            for col, team_id in enumerate(team_ids, 1):
                if team_id:
                    rows.append(self._public_record(match_id, path, sheet, row, year, "advertising", team_id, f"{market}:{label}", parse_money_wan(sheet.cell_value(row, col)) or 0.0))
        return rows

    def _parse_public_combined_reports(self, match_id: str, path: Path, sheet: xlrd.sheet.Sheet, year: int | None) -> list[dict[str, Any]]:
        team_ids = [clean_text(sheet.cell_value(1, col)) for col in range(2, sheet.ncols)]
        rows = []
        sections = [(2, 13, "comprehensive_expense"), (15, 26, "income_statement"), (28, 50, "balance_sheet")]
        for start, stop, section in sections:
            for row in range(start, min(stop, sheet.nrows)):
                metric = clean_text(sheet.cell_value(row, 1))
                if not metric or metric == "类型":
                    continue
                for col, team_id in enumerate(team_ids, 2):
                    value = parse_money_wan(sheet.cell_value(row, col))
                    if team_id and value is not None:
                        rows.append(self._public_record(match_id, path, sheet, row, year, section, team_id, metric, value))
        return rows

    def _parse_public_separate_report(self, match_id: str, path: Path, sheet: xlrd.sheet.Sheet, year: int | None, section: str) -> list[dict[str, Any]]:
        team_ids = [clean_text(sheet.cell_value(1, col)) for col in range(2, sheet.ncols)]
        rows = []
        start = 3 if section == "balance_sheet" and clean_text(sheet.cell_value(2, 1)) == "类型" else 2
        for row in range(start, sheet.nrows):
            metric = clean_text(sheet.cell_value(row, 1))
            if not metric:
                continue
            for col, team_id in enumerate(team_ids, 2):
                value = parse_money_wan(sheet.cell_value(row, col))
                if team_id and value is not None:
                    rows.append(self._public_record(match_id, path, sheet, row, year, section, team_id, metric, value))
        return rows

    def _parse_public_market_leader(self, match_id: str, path: Path, sheet: xlrd.sheet.Sheet, year: int | None) -> list[dict[str, Any]]:
        rows = []
        for row in range(1, sheet.nrows):
            values = [clean_text(sheet.cell_value(row, col)) for col in range(sheet.ncols)]
            values = [value for value in values if value]
            if values:
                rows.append({
                    "match_id": match_id,
                    "year": year,
                    "section": "market_leader",
                    "team_id": values[-1] if len(values) > 1 else None,
                    "metric": values[-2] if len(values) > 1 else values[0],
                    "value": values[-1] if len(values) > 1 else None,
                    "raw_values": values,
                    "source_path": path.relative_to(self.project_root).as_posix(),
                    "source_sheet": sheet.name,
                    "source_row": row + 1,
                    "provenance": "observed",
                })
        return rows

    def _parse_public_rows(self, match_id: str, path: Path, sheet: xlrd.sheet.Sheet, year: int | None, section: str) -> list[dict[str, Any]]:
        rows = []
        for row in range(sheet.nrows):
            values = [_json_value(sheet.cell_value(row, col)) for col in range(sheet.ncols)]
            if any(clean_text(value) for value in values):
                rows.append({
                    "match_id": match_id,
                    "year": year,
                    "section": section,
                    "team_id": None,
                    "metric": None,
                    "value": None,
                    "raw_values": values,
                    "source_path": path.relative_to(self.project_root).as_posix(),
                    "source_sheet": sheet.name,
                    "source_row": row + 1,
                    "provenance": "observed",
                })
        return rows

    def _public_record(self, match_id: str, path: Path, sheet: xlrd.sheet.Sheet, row: int, year: int | None, section: str, team_id: str, metric: str, value: Any) -> dict[str, Any]:
        return {
            "match_id": match_id,
            "year": year,
            "section": section,
            "team_id": team_id.upper(),
            "metric": metric,
            "value": value,
            "raw_values": None,
            "source_path": path.relative_to(self.project_root).as_posix(),
            "source_sheet": sheet.name,
            "source_row": row + 1,
            "provenance": "observed",
        }

    @staticmethod
    def _team_ids_from_public(files: list[Path]) -> set[str]:
        team_ids: set[str] = set()
        pattern = re.compile(r"^[A-Za-z]{2}\d{2}$")
        for path in files:
            workbook = _open_xls(path)
            for sheet in workbook.sheets():
                for row in range(min(3, sheet.nrows)):
                    for col in range(sheet.ncols):
                        value = clean_text(sheet.cell_value(row, col))
                        if value and pattern.match(value):
                            team_ids.add(value.upper())
        return team_ids

    def _parse_xa_global_orders(self, path: Path) -> list[dict[str, Any]]:
        sheet = load_workbook(path, data_only=True, read_only=True).active
        headers = [clean_text(value) for value in next(sheet.iter_rows(values_only=True))]
        rows = []
        for row_index, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
            data = {headers[index]: values[index] for index in range(min(len(headers), len(values)))}
            order_id = clean_text(data.get("订单编号"))
            if not order_id:
                continue
            owner = clean_text(data.get("所属用户"))
            status_key = next((header for header in headers if header and "状态" in header), None)
            rows.append({
                "match_id": "LX_XA",
                "order_id": order_id,
                "order_type": clean_text(data.get("类型")),
                "year": as_int(data.get("年份")),
                "market": clean_text(data.get("市场")),
                "product": clean_text(data.get("产品")),
                "quantity": as_number(data.get("数量")),
                "total_price_wan": parse_money_wan(data.get("总价")),
                "delivery_term_quarters": as_int(data.get("交货期")),
                "receivable_term_quarters": as_int(data.get("账期")),
                "iso": clean_text(data.get("ISO")) or "-",
                "owner_team_id": None if owner in {None, "-"} else owner.upper(),
                "status": clean_text(data.get(status_key)) if status_key else None,
                "delivered_period": None,
                "coverage_scope": "complete_global_order_pool",
                "source_path": path.relative_to(self.project_root).as_posix(),
                "source_sheet": sheet.title,
                "source_row": row_index,
                "provenance": "observed",
            })
        return rows

    @staticmethod
    def _reconstruct_order_pool(match_id: str, orders: list[dict[str, Any]]) -> list[dict[str, Any]]:
        unique = {}
        for order in orders:
            candidate = dict(order)
            candidate["coverage_scope"] = "observed_allocated_order"
            if candidate.get("order_type") is None:
                inferred_type = {"X": "选单", "J": "竞单", "B": "未知竞价类"}.get(candidate["order_id"][:1].upper())
                if inferred_type:
                    candidate["order_type"] = inferred_type
                    candidate["order_type_provenance"] = "inferred_from_identifier_prefix"
            unique.setdefault(candidate["order_id"], candidate)
        observed = list(unique.values())
        if not observed:
            return observed

        generated = []
        numeric_rows = [row for row in observed if row["order_id"].isdigit()]
        if numeric_rows:
            # The AB-family identifiers share the 180001–180487 universe across ten matches.
            universe = range(180001, 180488)
            by_number = {int(row["order_id"]): row for row in numeric_rows}
            templates = sorted(numeric_rows, key=lambda row: row["order_id"])
            for number in universe:
                if number in by_number:
                    continue
                template = templates[(number - 180001) % len(templates)]
                generated.append(MultiMatchDatasetBuilder._simulated_unassigned_order(match_id, str(number), template, "numeric_order_universe_180001_180487"))
        grouped: dict[str, dict[int, dict[str, Any]]] = defaultdict(dict)
        for row in observed:
            match = re.match(r"([A-Za-z]+)[^-]*-(\d+)$", row["order_id"])
            if match:
                grouped[match.group(1).upper()][int(match.group(2))] = row
        for kind, by_suffix in grouped.items():
            templates = [by_suffix[index] for index in sorted(by_suffix)]
            max_suffix = max(by_suffix)
            for suffix in range(1, max_suffix + 1):
                if suffix in by_suffix:
                    continue
                template = templates[(suffix - 1) % len(templates)]
                prefix_match = re.match(r"([A-Za-z]+[^-]*)-", template["order_id"])
                encoded_prefix = prefix_match.group(1) if prefix_match else kind
                order_id = f"{encoded_prefix}-{suffix:04d}"
                generated.append(MultiMatchDatasetBuilder._simulated_unassigned_order(match_id, order_id, template, f"observed_{kind}_suffix_range_1_{max_suffix}"))
        return observed + generated

    @staticmethod
    def _simulated_unassigned_order(match_id: str, order_id: str, template: dict[str, Any], assumption: str) -> dict[str, Any]:
        row = dict(template)
        row.update({
            "match_id": match_id,
            "order_id": order_id,
            "order_type": {"X": "选单", "J": "竞单", "B": "未知竞价类"}.get(order_id[:1].upper(), template.get("order_type")),
            "owner_team_id": None,
            "status": "模拟未分配",
            "delivered_period": None,
            "coverage_scope": "simulated_unassigned_order",
            "source_path": None,
            "source_sheet": None,
            "source_row": None,
            "simulation_assumption": assumption,
            "template_order_id": template["order_id"],
            "provenance": "simulated",
        })
        return row

    def _parse_xa_rules(self, path: Path) -> dict[str, Any]:
        sheet = load_workbook(path, data_only=True, read_only=True).active
        raw_rows = []
        for row_number, values in enumerate(sheet.iter_rows(values_only=True), 1):
            compact = [_json_value(value) for value in values if clean_text(value) is not None]
            raw_rows.append({"row": row_number, "values": compact})
        return {
            "match_id": "LX_XA",
            "rule_pack_id": "AAAAA_20160108_4.3.32.419",
            "binding_status": "confirmed_formal_source",
            "provenance": "observed",
            "source_path": path.relative_to(self.project_root).as_posix(),
            "parameters": {
                "initial_cash_wan": 675,
                "management_fee_per_quarter_wan": 14,
                "tax_rate": 0.25,
                "default_penalty_rate": 0.20,
                "bankruptcy": ["cash_flow_break", "negative_equity"],
                "first_year_has_orders": False,
                "long_loan": {"annual_rate": 0.12, "max_years": 4, "max_total_multiple_prior_equity": 3, "minimum_wan": 10, "application_timing": "year_start", "repayment_mode": "annual_interest_maturity_principal"},
                "short_loan": {"rate": 0.05, "application_timing": "quarter_start", "repayment_mode": "principal_and_interest_at_maturity"},
                "receivable_discount": {"terms_1_2": 0.08, "terms_3_4": 0.09, "mode": "joint"},
                "factories": {
                    "大厂房": {"purchase_wan": 481, "rent_wan_per_year": 51, "sale_wan": 481, "capacity": 6, "usage_limit": 4, "score": 11},
                    "中厂房": {"purchase_wan": 248, "rent_wan_per_year": 31, "sale_wan": 248, "capacity": 3, "usage_limit": 4, "score": 6},
                    "小厂房": {"purchase_wan": 72, "rent_wan_per_year": 9, "sale_wan": 72, "capacity": 1, "usage_limit": 4, "score": 2},
                },
                "production_lines": {
                    "手工线": {"investment_wan": 40, "investment_wan_per_quarter": 40, "install_quarters": 0, "production_quarters": 2, "conversion_wan_per_quarter": 10, "conversion_quarters": 0, "maintenance_wan_per_year": 15, "residual_value_wan": 10, "depreciation_fee_wan": 10, "depreciation_years": 4, "batch_capacity": 1, "score": 5},
                    "自动线": {"investment_wan": 140, "investment_wan_per_quarter": 70, "install_quarters": 2, "production_quarters": 1, "conversion_wan_per_quarter": 10, "conversion_quarters": 1, "maintenance_wan_per_year": 11, "residual_value_wan": 28, "depreciation_fee_wan": 28, "depreciation_years": 5, "batch_capacity": 1, "score": 8},
                    "柔性线": {"investment_wan": 180, "investment_wan_per_quarter": 60, "install_quarters": 3, "production_quarters": 1, "conversion_wan_per_quarter": 0, "conversion_quarters": 0, "maintenance_wan_per_year": 15, "residual_value_wan": 36, "depreciation_fee_wan": 36, "depreciation_years": 5, "batch_capacity": 1, "score": 10},
                    "租赁线": {"investment_wan": 0, "investment_wan_per_quarter": 0, "install_quarters": 0, "production_quarters": 1, "conversion_wan_per_quarter": 10, "conversion_quarters": 1, "maintenance_wan_per_year": 70, "residual_value_wan": -70, "depreciation_fee_wan": 0, "depreciation_years": 0, "batch_capacity": 1, "score": 0},
                },
                "markets": {
                    "本地": {"fee_wan_per_year": 8, "years": 1, "score": 6}, "区域": {"fee_wan_per_year": 8, "years": 1, "score": 6},
                    "国内": {"fee_wan_per_year": 10, "years": 1, "score": 7}, "亚洲": {"fee_wan_per_year": 12, "years": 3, "score": 10},
                    "国际": {"fee_wan_per_year": 16, "years": 3, "score": 14},
                },
                "iso": {"ISO9000": {"fee_wan_per_year": 21, "years": 2, "score": 6}, "ISO14000": {"fee_wan_per_year": 33, "years": 2, "score": 9}},
                "products": {
                    "P1": {"process_wan": 8, "development_wan_per_quarter": 14, "quarters": 1, "direct_cost_wan": 16, "score": 5, "bom": {"R1": 1}},
                    "P2": {"process_wan": 9, "development_wan_per_quarter": 14, "quarters": 1, "direct_cost_wan": 27, "score": 7, "bom": {"R2": 1, "R3": 1}},
                    "P3": {"process_wan": 10, "development_wan_per_quarter": 15, "quarters": 2, "direct_cost_wan": 36, "score": 9, "bom": {"R1": 1, "R3": 2}},
                    "P4": {"process_wan": 11, "development_wan_per_quarter": 16, "quarters": 3, "direct_cost_wan": 48, "score": 12, "bom": {"P2": 1, "R4": 1}},
                    "P5": {"process_wan": 11, "development_wan_per_quarter": 25, "quarters": 4, "direct_cost_wan": 56, "score": 14, "bom": {"P3": 1, "R3": 1}},
                },
                "materials": {"R1": {"price_wan": 8, "lead_quarters": 1}, "R2": {"price_wan": 9, "lead_quarters": 2}, "R3": {"price_wan": 9, "lead_quarters": 1}, "R4": {"price_wan": 10, "lead_quarters": 2}},
                "market_iso_payment_timing": "year_end",
                "product_development_payment_timing": "quarter_end",
                "production_line_depreciation_timing": "not_in_completion_year",
                "factory_sale_receivable_term_quarters": 4,
                "emergency_material_price_multiplier": 2,
                "emergency_product_price_multiplier": 3,
                "max_factory_count": 4,
                "minimum_order_advertising_wan": 10,
                "rounding": {"default_penalty": "half_up", "inventory_auction": "floor", "receivable_discount_fee": "ceil", "income_tax": "half_up", "loan_interest": "half_up"},
                "selection_priority": ["prior_market_leader", "market_product_advertising", "market_total_advertising", "market_sales_rank", "advertising_submission_time"],
                "score_formula": "score = owner_equity * (1 + development_potential / 100)",
            },
            "raw_rule_rows": raw_rows,
        }

    def _build_old_rules(self, match_id: str, events: list[dict[str, Any]], enterprise_files: list[Path], global_orders: list[dict[str, Any]]) -> dict[str, Any]:
        candidate = json.loads(self.old_candidate_rules.read_text(encoding="utf-8"))
        capital = sorted({row["amount_wan"] for row in events if row["action"] == "capital_injection" and row["amount_wan"] is not None})
        fees = sorted({abs(row["amount_wan"]) for row in events if row["action"] == "administrative_fee" and row["amount_wan"] is not None})
        observed_actions = sorted({row["action"] for row in events if row["action"]})
        observed_order_types = sorted({row.get("order_type") for row in global_orders if row.get("order_type") and row.get("provenance") == "observed"})
        simulated_orders = sum(row.get("provenance") == "simulated" for row in global_orders)
        # Candidate values are retained as a reference only. A match-specific pack
        # records which parameters are supported by its own event/report evidence.
        inferred_parameters = {
            "initial_cash_wan_observed": sorted({row["balance_wan"] for row in events if row["action"] == "capital_injection" and row["balance_wan"] is not None}),
            "management_fee_per_quarter_wan_observed": fees,
            "observed_action_set": observed_actions,
            "observed_order_types": observed_order_types,
            "competitions_with_explicit_order_type": bool(observed_order_types),
        }
        inherited_from_xa = [
            "cash_transition_identity",
            "quarterly_event_ordering_interface",
            "bankruptcy_state_fields",
            "report_statement_names",
            "provenance_and_coverage_contract",
        ]
        parameters = self._candidate_parameters_wan(candidate)
        if capital:
            parameters["initial_cash_wan"] = capital[0]
        if fees:
            parameters["management_fee_per_quarter_wan"] = fees[0]
        return {
            "match_id": match_id,
            "rule_pack_id": f"{match_id}_historical_rule_unknown",
            "binding_status": "reconstructed_match_specific",
            "provenance": "inferred",
            "source_scope": {"enterprise_file_count": len(enterprise_files), "global_order_record_count": len(global_orders)},
            "observed_fingerprint": {"initial_capital_amounts_wan": capital, "management_fee_amounts_wan": fees, "action_count": len(observed_actions)},
            "inferred_parameters": inferred_parameters,
            "parameters": parameters,
            "parameter_provenance": {
                "initial_cash_wan": "observed_event_fingerprint",
                "management_fee_per_quarter_wan": "observed_event_fingerprint",
                "factory_line_market_iso_product_material_parameters": "inferred_from_600W_candidate_pack_and_checked_against_exports",
                "bankruptcy_selection_and_scoring_semantics": "inherited_from_XA_when_not_contradicted",
            },
            "inherited_xa_interfaces_only": inherited_from_xa,
            "simulation_policy": {
                "unallocated_orders_generated": simulated_orders > 0,
                "simulated_unallocated_order_count": simulated_orders,
                "generate_competition_order_types_only_if_observed": True,
                "do_not_generate_when_complete_pool_observed": True,
                "warning": "Generated rows are scenario placeholders, not official historical facts.",
            },
            "candidate_rule_pack": {
                "source_path": self.old_candidate_rules.relative_to(self.project_root).as_posix(),
                "binding_status": "unconfirmed_not_auto_bound_to_match",
                "currency_unit": "yuan",
                "data": candidate,
            },
            "common_xa_rule_domains_available_for_reuse": [
                "production_line", "financing", "factory", "market", "iso", "product", "materials", "bankruptcy", "scoring",
            ],
            "known_gaps": [
                "No formal rule workbook is bound to this match.",
                "The complete pre-selection global order pool is absent from the source export.",
                "Allocated orders are observed; unallocated rows, when generated, are simulated from observed identifier ranges.",
            ],
        }

    @staticmethod
    def _candidate_parameters_wan(candidate: dict[str, Any]) -> dict[str, Any]:
        to_wan = lambda value: value / 10000 if value is not None else None
        return {
            "initial_cash_wan": to_wan(candidate["meta"]["initial_capital"]),
            "management_fee_per_quarter_wan": to_wan(candidate["meta"]["management_fee_per_quarter"]),
            "tax_rate": 0.25,
            "default_penalty_rate": 0.20,
            "bankruptcy": ["cash_flow_break", "negative_equity"],
            "first_year_has_orders": False,
            "long_loan": {
                "annual_rate": candidate["loans"]["long_term_rate"],
                "max_years": candidate["loans"]["long_term_max_years"],
                "max_total_multiple_prior_equity": 3,
            },
            "short_loan": {"rate": candidate["loans"]["short_term_rate"]},
            "receivable_discount": {
                "terms_1_2": candidate["loans"]["discount_rates"]["q1_q2"],
                "terms_3_4": candidate["loans"]["discount_rates"]["q3_q4"],
            },
            "factories": {
                row["name"]: {
                    "purchase_wan": to_wan(row["purchase_price"]),
                    "rent_wan_per_year": to_wan(row["rent_price_per_year"]),
                    "capacity": row["line_capacity"],
                    "score": row["score"],
                }
                for row in candidate["factories"]
            },
            "production_lines": {
                row["name"]: {
                    "investment_wan_per_quarter": to_wan(row["investment_per_quarter"]),
                    "install_quarters": row["install_periods"],
                    "production_quarters": row["production_periods"],
                    "maintenance_wan_per_year": to_wan(row["maintenance_per_year"]),
                    "score": row["score"],
                }
                for row in candidate["lines"]
            },
            "markets": {
                row["name"]: {"fee_wan_per_year": to_wan(row["development_fee_per_year"]), "years": row["development_years"], "score": row["score"]}
                for row in candidate["markets"]
            },
            "iso": {
                row["name"]: {"fee_wan_per_year": to_wan(row["development_fee_per_year"]), "years": row["development_years"], "score": row["score"]}
                for row in candidate["iso"]
            },
            "products": {
                row["name"]: {
                    "process_wan": to_wan(row["process_fee"]),
                    "development_wan_per_quarter": to_wan(row["rd_fee_per_quarter"]),
                    "quarters": row["rd_periods"],
                    "direct_cost_wan": to_wan(row["direct_cost"]),
                    "score": row["score"],
                }
                for row in candidate["products"]
            },
            "materials": {
                row["name"]: {"price_wan": to_wan(row["unit_price"]), "lead_quarters": row["lead_time_quarters"]}
                for row in candidate["materials"]
            },
            "bom": candidate["bom"],
            "selection_priority": ["prior_market_leader", "market_product_advertising", "market_total_advertising", "market_sales_rank", "submission_time"],
            "score_formula": "score = owner_equity * (1 + development_potential / 100)",
        }

    @staticmethod
    def _parse_score_assets(team_id: str, workbook: xlrd.book.Book) -> dict[str, Any]:
        result = {"markets": [], "products": [], "iso": [], "purchased_factories": [], "completed_lines": []}
        if "研发认证" in workbook.sheet_names():
            sheet = workbook.sheet_by_name("研发认证")
            for kind, start_col in (("markets", 1), ("products", 7), ("iso", 13)):
                for row in range(3, sheet.nrows):
                    name = clean_text(sheet.cell_value(row, start_col)) if start_col < sheet.ncols else None
                    remaining = clean_text(sheet.cell_value(row, start_col + 3)) if start_col + 3 < sheet.ncols else None
                    if name and remaining == "-":
                        result[kind].append(name)
        if "厂房与生产线" in workbook.sheet_names():
            sheet = workbook.sheet_by_name("厂房与生产线")
            line_title = next((row for row in range(sheet.nrows) if clean_text(sheet.cell_value(row, 1)) == "生产线信息"), sheet.nrows)
            for row in range(3, line_title):
                if clean_text(sheet.cell_value(row, 3)) == "购买":
                    name = clean_text(sheet.cell_value(row, 2))
                    if name:
                        result["purchased_factories"].append(name)
            for row in range(line_title + 2, sheet.nrows):
                name = clean_text(sheet.cell_value(row, 2))
                completion = clean_text(sheet.cell_value(row, 10)) if sheet.ncols > 10 else None
                if name and completion and completion != "-":
                    result["completed_lines"].append(name)
        return result

    @staticmethod
    def _parse_xa_bankruptcies(path: Path) -> dict[str, str]:
        text = "\n".join(paragraph.text for paragraph in Document(path).paragraphs)
        result = {}
        for team_id, year, quarter in re.findall(r"(XA\d{2})组第(\d+)年第(\d+)季破产", text):
            result[team_id] = f"Y{year}Q{quarter}"
        return result

    def _build_quarter_states(self, match_id: str, teams: list[dict[str, Any]], events: list[dict[str, Any]], bankruptcies: dict[str, str]) -> list[dict[str, Any]]:
        by_team: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for event in events:
            if event["included_in_match"]:
                by_team[event["team_id"]].append(event)
        rows = []
        for team in teams:
            team_id = team["team_id"]
            team_events = sorted(by_team[team_id], key=lambda row: row["sequence_in_source"])
            period_events: dict[tuple[int, int], list[dict[str, Any]]] = defaultdict(list)
            for event in team_events:
                period_events[(event["year"], event["quarter"])].append(event)
            previous = None
            bankruptcy_index = None
            if team_id in bankruptcies:
                y, q = parse_period(bankruptcies[team_id])
                bankruptcy_index = (y - 1) * 4 + q if y and q else None
            for year in OPERATING_YEARS:
                for quarter in OPERATING_QUARTERS:
                    period_index = (year - 1) * 4 + quarter
                    current = period_events[(year, quarter)]
                    start_cash = previous
                    end_cash = current[-1]["balance_wan"] if current else previous
                    if current and start_cash is None:
                        first = current[0]
                        start_cash = (first["balance_wan"] - first["amount_wan"]) if first["balance_wan"] is not None and first["amount_wan"] is not None else None
                    if current:
                        status = "derived_from_observed_events"
                    elif bankruptcy_index is not None and period_index > bankruptcy_index:
                        status = "not_applicable_after_bankruptcy"
                    elif previous is not None:
                        status = "derived_carry_forward"
                    else:
                        status = "missing"
                    rows.append({
                        "match_id": match_id,
                        "state_id": f"{match_id}:{team_id}:Y{year}Q{quarter}",
                        "team_id": team_id,
                        "period_index": period_index,
                        "period": f"Y{year}Q{quarter}",
                        "start_cash_wan": start_cash,
                        "end_cash_wan": end_cash,
                        "event_count": len(current),
                        "status": status,
                        "provenance": "derived" if status.startswith("derived") else ("missing" if status == "missing" else "derived"),
                    })
                    previous = end_cash
        return rows

    def _build_final_states(self, match_id: str, teams: list[dict[str, Any]], events: list[dict[str, Any]], reports: list[dict[str, Any]], assets: dict[str, dict[str, Any]], rules: dict[str, Any], bankruptcies: dict[str, str]) -> list[dict[str, Any]]:
        event_cash = {}
        for event in sorted(events, key=lambda row: (row["team_id"], row["sequence_in_source"])):
            if event["included_in_match"]:
                event_cash[event["team_id"]] = event["balance_wan"]
        equity = {}
        for report in reports:
            if report["year"] == 5 and report["metric"] in {"所有者权益合计", "所有者权益"} and report["report_variant"] in {"system", "系统"}:
                equity[report["team_id"]] = report["value_wan"]
        params = rules.get("parameters", {})
        score_maps = {key: {name: spec.get("score", 0) for name, spec in params.get(key, {}).items()} for key in ("markets", "products", "iso", "production_lines", "factories")}
        rows = []
        for team in teams:
            team_id = team["team_id"]
            team_assets = assets.get(team_id, {"markets": [], "products": [], "iso": [], "purchased_factories": [], "completed_lines": []})
            potential = None
            if any(score_maps.values()):
                potential = sum(score_maps["markets"].get(name, 0) for name in team_assets["markets"])
                potential += sum(score_maps["products"].get(name, 0) for name in team_assets["products"])
                potential += sum(score_maps["iso"].get(name, 0) for name in team_assets["iso"])
                potential += sum(score_maps["factories"].get(name, 0) for name in team_assets["purchased_factories"])
                potential += sum(score_maps["production_lines"].get(name, 0) for name in team_assets["completed_lines"])
            owner_equity = equity.get(team_id)
            score = owner_equity * (1 + potential / 100) if owner_equity is not None and potential is not None else None
            rows.append({
                "match_id": match_id,
                "team_id": team_id,
                "official_close_period": bankruptcies.get(team_id, "Y5Q4"),
                "final_cash_wan": event_cash.get(team_id),
                "export_cash_wan": team["export_cash_wan"],
                "owner_equity_wan": owner_equity,
                "development_potential": potential,
                "score": score,
                "score_provenance": "derived" if score is not None else "missing",
                "bankruptcy_period": bankruptcies.get(team_id),
                "assets": team_assets,
                "provenance": "derived",
            })
        return rows

    @staticmethod
    def _build_results(match_id: str, final_states: list[dict[str, Any]], bankruptcies: dict[str, str], rules: dict[str, Any]) -> dict[str, Any]:
        ranked = [row for row in final_states if row["score"] is not None and row["team_id"] not in bankruptcies]
        ranked.sort(key=lambda row: (-row["score"], row["team_id"]))
        official = {team_id: score for team_id, score in XA_OFFICIAL_RANKING} if match_id == "LX_XA" else {}
        ranking = [{
            "rank": index,
            "team_id": row["team_id"],
            "official_score": official.get(row["team_id"]),
            "recomputed_score": row["score"],
            "rounded_recomputed_score": math.floor(row["score"] + 0.5),
            "rounding_rule": "half_up_for_nonnegative_score",
            "official_score_matches": official.get(row["team_id"]) == math.floor(row["score"] + 0.5) if row["team_id"] in official else None,
            "owner_equity_wan": row["owner_equity_wan"],
            "development_potential": row["development_potential"],
            "provenance": {"official_score": "observed", "recomputed_score": "derived"} if row["team_id"] in official else {"official_score": "missing", "recomputed_score": "derived"},
        } for index, row in enumerate(ranked, 1)]
        return {
            "match_id": match_id,
            "ranking": ranking,
            "ranking_status": "derived_and_checked_against_official_score" if match_id == "LX_XA" else "simulated_ranking_from_reconstructed_rule_pack",
            "bankruptcies": [{"team_id": team_id, "period": period, "provenance": "observed"} for team_id, period in sorted(bankruptcies.items())],
            "rule_pack_id": rules["rule_pack_id"],
        }

    def _quality(self, match_id: str, teams: list[dict[str, Any]], events: list[dict[str, Any]], team_orders: list[dict[str, Any]], global_orders: list[dict[str, Any]], reports: list[dict[str, Any]], quarter_states: list[dict[str, Any]], sources: list[Path]) -> dict[str, Any]:
        continuity_errors = []
        by_team = defaultdict(list)
        for event in events:
            if event["included_in_match"]:
                by_team[event["team_id"]].append(event)
        for team_id, rows in by_team.items():
            previous = None
            for event in sorted(rows, key=lambda row: row["sequence_in_source"]):
                if previous is not None and event["amount_wan"] is not None and event["balance_wan"] is not None:
                    expected = previous + event["amount_wan"]
                    if abs(expected - event["balance_wan"]) > 1e-6:
                        continuity_errors.append({"event_id": event["event_id"], "expected": expected, "actual": event["balance_wan"]})
                previous = event["balance_wan"]
        global_ids = [row["order_id"] for row in global_orders]
        duplicate_orders = [order_id for order_id, count in Counter(global_ids).items() if count > 1]
        team_order_ids = {row["order_id"] for row in team_orders}
        global_order_ids = set(global_ids)
        unmatched_team_orders = sorted(team_order_ids - global_order_ids)
        global_by_id = {row["order_id"]: row for row in global_orders}
        order_field_mismatches = []
        for order in team_orders:
            global_order = global_by_id.get(order["order_id"])
            if not global_order:
                continue
            for field in ("owner_team_id", "market", "product", "quantity", "total_price_wan"):
                if order[field] != global_order[field]:
                    order_field_mismatches.append({"order_id": order["order_id"], "field": field, "team_value": order[field], "global_value": global_order[field]})
        hashes = [sha256_file(path) for path in sources]
        unmapped_operating_actions = [event["event_id"] for event in events if event["included_in_match"] and event["action"] is None]
        return {
            "match_id": match_id,
            "record_counts": {
                "teams": len(teams), "events": len(events), "operating_events": sum(row["included_in_match"] for row in events),
                "team_orders": len(team_orders), "global_orders": len(global_orders),
                "observed_global_orders": sum(row.get("provenance") == "observed" for row in global_orders),
                "simulated_global_orders": sum(row.get("provenance") == "simulated" for row in global_orders),
                "unassigned_global_orders": sum(row.get("owner_team_id") is None for row in global_orders),
                "reports": len(reports), "quarter_states": len(quarter_states),
            },
            "coverage": {
                "schema": "complete",
                "enterprise_exports": "observed",
                "event_stream": "observed_from_cash_flow_sheet",
                "quarter_cash_states": "derived_from_observed_events",
                "global_orders": "complete_observed_pool" if match_id == "LX_XA" else "observed_allocated_plus_simulated_unassigned",
                "formal_rules": "complete_observed" if match_id == "LX_XA" else "reconstructed_match_specific_for_simulation",
                "final_ranking": "derived_from_formal_rule" if match_id == "LX_XA" else "simulated_from_reconstructed_rule_pack",
            },
            "checks": {
                "cash_continuity": {"passed": not continuity_errors, "error_count": len(continuity_errors), "errors": continuity_errors[:20]},
                "operating_action_mapping": {"passed": not unmapped_operating_actions, "unmapped_count": len(unmapped_operating_actions), "event_ids": unmapped_operating_actions[:50]},
                "global_order_unique": {"passed": not duplicate_orders, "duplicates": duplicate_orders},
                "team_orders_present_in_global_view": {"passed": not unmatched_team_orders, "unmatched": unmatched_team_orders[:50]},
                "team_global_order_fields": {"passed": not order_field_mismatches, "mismatch_count": len(order_field_mismatches), "mismatches": order_field_mismatches[:50]},
                "quarter_state_shape": {"passed": len(quarter_states) == len(teams) * 20, "expected": len(teams) * 20, "actual": len(quarter_states)},
                "duplicate_source_hashes": {"passed": len(hashes) == len(set(hashes)), "duplicate_count": len(hashes) - len(set(hashes))},
            },
        }

    def _write_readme(self, catalog: dict[str, Any]) -> None:
        rows = [
            "# GoAI 中文企业经营比赛统一数据集 v2",
            "",
            "本目录将 13 场历史导出和 1 场省赛完整资料统一为同一结构。所有赛场都具有相同文件名和字段结构；内容完整度通过 `provenance`、`coverage_scope`、`observed_complete` 和 `quality.json` 明确表达。旧比赛缺少正式规则原件和赛前全局订单池，现已生成面向仿真的逐场重建规则集与未分配订单；这些记录明确标记为 `inferred` 或 `simulated`，不会冒充官方事实。",
            "",
            "## 企业事件流",
            "",
            "主事件流来自每个企业工作簿的 `现金流量表` sheet，第 4 行开始依次为 `ID、动作、资金、余额、时间、备注`。`events.jsonl` 是其标准化结果；`订单信息`、`厂房与生产线`、`研发认证`、`库存信息` 和三张报表是事件结果的状态与审计证据。",
            "",
            "## 每场比赛目录",
            "",
            "| 文件 | 用途 |",
            "| --- | --- |",
            "| `manifest.json` | 比赛身份、来源文件、哈希和完整性声明 |",
            "| `rules.json` | 正式规则、观察指纹、候选规则和缺口 |",
            "| `teams.jsonl` | 企业元数据及导出状态 |",
            "| `events.jsonl` | 企业经营事件流 |",
            "| `global_orders.jsonl` | XA 完整观测订单池；旧场次为已分配订单加模拟未分配订单 |",
            "| `annual_public.jsonl` | 年度广告、公共报表、市场老大和生产线巡盘 |",
            "| `reports.jsonl` | 企业年度报表和企业广告 |",
            "| `final_states.jsonl` | 比赛结束或破产时状态与评分输入 |",
            "| `quarter_states.jsonl` | 每队固定 20 个季度现金切片 |",
            "| `results.json` | 排名、评分与破产时间 |",
            "| `raw_cells.jsonl` | 全部 XLS 非空单元格保真长表 |",
            "| `quality.json` | 现金连续性、订单、覆盖率和形状检查 |",
            "",
            "## 赛场概览",
            "",
            "| match_id | 企业数 | 经营事件 | 订单总数 | 模拟订单 | 订单覆盖 | 观测完整 |",
            "| --- | ---: | ---: | ---: | ---: | --- | --- |",
        ]
        for match in catalog["matches"]:
            rows.append(f"| {match['match_id']} | {match['team_count']} | {match['operating_event_count']} | {match['global_order_count']} | {match['simulated_order_count']} | {match['global_order_coverage']} | {str(match['observed_complete']).lower()} |")
        rows.extend([
            "",
            "## 使用顺序",
            "",
            "先用 XA 建立可审计的规则状态机和报表重放基准，再用旧赛场训练动作识别、异常检测、VPD/PSS 指标和行为模式。旧场次的模拟未分配订单只适合训练接口、压力测试和候选策略，不可用于评估历史参赛队当时真实可见的机会集合。",
            "",
        ])
        (self.output_root / "README.md").write_text("\n".join(rows), encoding="utf-8")
