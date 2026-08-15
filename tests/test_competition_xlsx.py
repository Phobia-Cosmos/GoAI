import json
import zipfile
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

from goai_data.competition_xlsx import ENTERPRISE_SHEETS, SimulatedCompetitionXlsxImporter, _write_enterprise_workbook, build_competition_xlsx_archive, export_competition_xlsx
from goai_data.full_sandbox import FullCompetitionArena, FullFinancialDynamics, SeededHeuristicPolicy, generate_global_orders, generate_simulated_rule_pack


DATASET_ROOT = Path("/home/undefined/Disk/datasets/goai/processed/v2/matches")


def load_base(match_id: str) -> dict:
    match_dir = DATASET_ROOT / match_id
    path = match_dir / "rules_inferred_v2.json"
    if not path.exists():
        path = match_dir / "rules.json"
    return json.loads(path.read_text(encoding="utf-8"))


def run_small_match(base_match: str, seed: int = 100) -> tuple[dict, list[dict], FullCompetitionArena]:
    rules = generate_simulated_rule_pack(load_base(base_match), seed=seed, match_id=f"SIM_{base_match}", team_count=2, source_match_id=base_match)
    orders = generate_global_orders(rules, seed=seed + 1, orders_per_year=2)
    dynamics = FullFinancialDynamics(rules)
    arena = FullCompetitionArena(dynamics, rules["participants"]["team_ids"], orders)
    policies = {team_id: SeededHeuristicPolicy(team_id, seed) for team_id in arena.agent_ids}
    observations = arena.reset(seed=seed)
    while not arena.terminated:
        result = arena.step({team_id: policies[team_id].act(observations[team_id]) for team_id in arena.agent_ids})
        observations = result.observations
    return rules, orders, arena


def read_jsonl(path: Path) -> list[dict]:
    return [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line]


def test_every_normalized_match_can_be_used_as_a_random_generator_template() -> None:
    match_ids = sorted(path.name for path in DATASET_ROOT.iterdir() if path.is_dir())
    assert set(match_ids) >= {"AB", "AG", "CA", "CB", "CD", "CE", "EA", "EB", "EC", "EF", "LX_XA", "OP", "ZY", "ZZ"}
    for index, match_id in enumerate(match_ids):
        rules = generate_simulated_rule_pack(load_base(match_id), seed=200 + index, match_id=f"SIM_{match_id}", team_count=3, source_match_id=match_id)
        orders = generate_global_orders(rules, seed=300 + index, orders_per_year=1)
        state = FullFinancialDynamics(rules).initial_state(rules["participants"]["team_ids"][0], orders=orders)
        assert rules["generation"]["source_match_id"] == match_id
        assert rules["parent_rule_pack_id"] == load_base(match_id).get("rule_pack_id")
        assert rules["provenance"] == "simulated"
        assert len(orders) == 4
        assert abs(state.balance_gap_wan) <= 1e-6


@pytest.mark.parametrize("base_match", ["AB", "AG", "LX_XA"])
def test_competition_xlsx_round_trip_uses_visible_competition_tables(tmp_path: Path, base_match: str) -> None:
    rules, orders, arena = run_small_match(base_match)
    bundle = tmp_path / "bundle"
    manifest = export_competition_xlsx(bundle, rules=rules, orders=orders, arena=arena)
    assert manifest["team_count"] == 2
    assert manifest["order_count"] == len(orders)
    assert manifest["provenance"] == "simulated"
    assert manifest["files"]["annual_public"] == [f"{year}.xlsx" for year in range(1, 7)]
    annual = load_workbook(bundle / "1.xlsx", data_only=True)
    assert tuple(name for name in annual.sheetnames if not name.startswith("_")) == (
        "第1年广告投放",
        "第1年广告投放(格式二)",
        "第1年三张报表",
        "生产线信息",
        "第1年市场老大",
    )

    for relative in manifest["files"]["enterprise"]:
        workbook = load_workbook(bundle / relative, data_only=True)
        assert tuple(name for name in workbook.sheetnames if not name.startswith("_")) == ENTERPRISE_SHEETS
        assert workbook["现金流量表"]["C3"].value == "动作"
        assert workbook["订单信息"]["B3"].value == "订单编号"
        assert workbook["三张报表"]["B2"].value == "年度"

    imported_dir = tmp_path / "imported"
    imported = SimulatedCompetitionXlsxImporter(bundle, imported_dir).import_bundle()
    teams = read_jsonl(imported_dir / "teams.jsonl")
    imported_orders = read_jsonl(imported_dir / "global_orders.jsonl")
    events = read_jsonl(imported_dir / "events.jsonl")
    reports = read_jsonl(imported_dir / "reports.jsonl")
    assert imported["counts"]["teams"] == len(arena.states)
    assert len(imported_orders) == len(orders)
    # 与真实“初始元年 + 第1年至第5年”六列布局一致。
    assert len(reports) == len(arena.states) * 6 * (len(range(3, 14)) + len(range(16, 27)) + len(range(30, 51)))
    assert all(row["provenance"] == "simulated" for row in teams + imported_orders + events + reports)

    final_cash = {team_id: state.cash_wan for team_id, state in arena.states.items()}
    imported_cash = {row["team_id"]: row["export_cash_wan"] for row in teams}
    assert imported_cash == pytest.approx(final_cash)
    by_team = {team_id: [] for team_id in final_cash}
    for event in events:
        by_team[event["team_id"]].append(event)
    for team_id, rows in by_team.items():
        rows.sort(key=lambda row: row["sequence_in_source"])
        assert rows[-1]["balance_wan"] == pytest.approx(final_cash[team_id])
        for previous, current in zip(rows, rows[1:]):
            assert previous["balance_wan"] + current["amount_wan"] == pytest.approx(current["balance_wan"])


def test_competition_archive_matches_the_reference_directory_family() -> None:
    rules, orders, arena = run_small_match("LX_XA", seed=509)
    payload = build_competition_xlsx_archive(rules=rules, orders=orders, arena=arena)
    match_id = rules["match_id"]
    with zipfile.ZipFile(BytesIO(payload)) as archive:
        names = set(archive.namelist())
        assert {f"{match_id}/{year}.xlsx" for year in range(1, 7)} <= names
        assert {f"{match_id}/{team_id}.xlsx" for team_id in arena.agent_ids} <= names
        assert f"{match_id}比赛规则.xlsx" in names
        assert f"{match_id}订单详情.xlsx" in names
        assert f"{match_id}最终排名和破产信息.xlsx" in names
        assert {"manifest.json", "导出说明.txt"} <= names
        rules_book = load_workbook(BytesIO(archive.read(f"{match_id}比赛规则.xlsx")), data_only=True)
        assert rules_book["Sheet1"]["A5"].value == "名称"
        assert rules_book["Sheet1"]["L5"].value == "分值"
        enterprise_book = load_workbook(BytesIO(archive.read(f"{match_id}/{arena.agent_ids[0]}.xlsx")), data_only=True)
        assert tuple(name for name in enterprise_book.sheetnames if not name.startswith("_")) == ENTERPRISE_SHEETS
        results_book = load_workbook(BytesIO(archive.read(f"{match_id}最终排名和破产信息.xlsx")), data_only=True)
        assert tuple(name for name in results_book.sheetnames if not name.startswith("_")) == ("最终排名", "破产信息", "全部企业终局")
        assert results_book["全部企业终局"].max_row == len(arena.agent_ids) + 1


def test_enterprise_export_preserves_timestamps_and_uses_chinese_cashflow_notes(tmp_path: Path) -> None:
    rules = generate_simulated_rule_pack(load_base("LX_XA"), seed=817, match_id="SIM_FIELDS", team_count=2, source_match_id="LX_XA")
    dynamics = FullFinancialDynamics(rules)
    state = dynamics.initial_state("SIM_FIELDS01")
    for action in (
        {"action_type": "long_loan_borrow", "parameters": {"principal_wan": 100, "term_years": 4}},
        {"action_type": "develop_product", "parameters": {"target": "P2"}},
        {"action_type": "buy_workshop", "parameters": {"factory": "小厂房"}},
        {"action_type": "buy_product_line", "parameters": {"line_type": "自动线", "product_id": "P1"}},
        {"action_type": "material_order", "parameters": {"materials": {"R1": 2}}},
    ):
        transition = dynamics.apply(state, action)
        assert transition.status == "success"
        state = transition.state

    assert state.completed_development[0]["completed_period"] == "Y1Q1"
    workbook_path = tmp_path / "SIM_FIELDS01.xlsx"
    _write_enterprise_workbook(workbook_path, state, rules)
    workbook = load_workbook(workbook_path, data_only=True)

    loans = workbook["银行贷款"]
    assert loans["G4"].value == "第1年1季"
    development = workbook["研发认证"]
    assert development["K5"].value == "第1年1季"
    assert development["M5"].value == "第1年1季"
    assets = workbook["厂房与生产线"]
    assert assets["J4"].value == "第1年1季"
    assert assets["K8"].value == "第1年3季"
    assert assets["L8"].value == "第1年1季"
    notes = [workbook["现金流量表"].cell(row, 7).value for row in range(4, workbook["现金流量表"].max_row + 1)]
    assert any(note == "申请长期贷款" for note in notes)
    assert any(note.startswith("订购生产线[自动线]") for note in notes)
    assert all("{" not in str(note) and "}" not in str(note) for note in notes)
